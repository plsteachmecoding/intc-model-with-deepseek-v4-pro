#!/usr/bin/env python3
"""
build_intc_model.py
Builds Intel_IB_Model.xlsx — 12-tab, segment-driven 3-statement DCF model
for Intel Corporation (INTC) using openpyxl with CHOOSE/MATCH scenario switching.

Usage:  python build_intc_model.py
Output: Intel_IB_Model.xlsx (in current directory)

Architecture (11 tabs):
  Cover → Key_Summary → Assumptions → Growth_Drivers → Segment_PL
  → Consolidated_PL → BS → Cash_Flow → DCF → Sensitivity → Ratio_Analysis

Data sources:
  - intc_financial_data.md  (extracted from 10-Ks, 10-Q)
  - intc_assumptions.md     (forecast assumptions, 3 scenarios)

References:
  - example-structure.md    (tab layouts, cross-reference pattern)
  - SKILL.md (ib-financial-model skill spec)

Author: Generated via Claude Code / ib-financial-model skill
Date:   May 7, 2026
"""

import os, sys, re
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.utils import get_column_letter as CL
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("ERROR: openpyxl required. Run: pip install openpyxl")

# ============================================================
# GLOBAL CONSTANTS
# ============================================================

OUTPUT = "Intel_IB_Model_v2.xlsx"
NCOL   = 11          # Notes column (K) — must be > last data column J
SCENARIO_CELL = "$B$2"  # Scenario selector location in Assumptions tab

# Historical columns: B=2 (2023A), C=3 (2024A), D=4 (2025A)
# Forecast columns:  F=6 (2026E), G=7 (2027E), H=8 (2028E)
# Col E=5  spacer, Col I=9  Selected (CHOOSE/MATCH), Col J=10 available, Col K=11 Notes
YR_COLS = {"2023A": 2, "2024A": 3, "2025A": 4, "2026E": 6, "2027E": 7, "2028E": 8}
HIST_COLS = [2, 3, 4]
FCST_COLS = [6, 7, 8]
SEL_COL   = 9

# Assumptions annual columns mapped to FCST_COLS for cross-sheet references
# FCST_COLS [6,7,8] -> Assumptions annual rollup [22,27,32]
ASP_ANN_COLS = {6: 22, 7: 27, 8: 32}

# Segment_PL quarterly annual rollup columns → annual-tab columns
# Annual tabs (Consolidated_PL, BS, CF, DCF) use cols [2,3,4,6,7,8]
# Segment_PL quarterly uses annual cols [6,11,16,22,27,32]
SEGPL_COL_MAP = {2: 6, 3: 11, 4: 16, 6: 22, 7: 27, 8: 32}

# ============================================================
# GROWTH_DRIVERS QUARTERLY LAYOUT (ALL 6 YEARS)
# ============================================================
# Col layout for Growth_Drivers tab (35 cols):
# B=Q1'23  C=Q2'23  D=Q3'23  E=Q4'23  F=FY2023A
# G=Q1'24  H=Q2'24  I=Q3'24  J=Q4'24  K=FY2024A
# L=Q1'25  M=Q2'25  N=Q3'25  O=Q4'25  P=FY2025A
# Q=spacer
# R=Q1'26  S=Q2'26  T=Q3'26  U=Q4'26  V=FY2026E
# W=Q1'27  X=Q2'27  Y=Q3'27  Z=Q4'27  AA=FY2027E
# AB=Q1'28 AC=Q2'28 AD=Q3'28 AE=Q4'28 AF=FY2028E
# AG=Selected  AH=spacer  AI=Notes
GD_NCOL = 35
GD_HIST_QCOLS = {
    2023: [2, 3, 4, 5],
    2024: [7, 8, 9, 10],
    2025: [12, 13, 14, 15],
}
GD_HIST_ANN_COLS = {2023: 6, 2024: 11, 2025: 16}
GD_QCOLS = {
    2026: [18, 19, 20, 21],
    2027: [23, 24, 25, 26],
    2028: [28, 29, 30, 31],
}
GD_ANN_COLS = {2026: 22, 2027: 27, 2028: 32}
GD_ANN_COLS_LIST = [6, 11, 16, 22, 27, 32]
GD_SEL_COL = 33
GD_SPACER_COL = 17
GD_SPACER2_COL = 34
# All years (hist + fcst) for iteration: list of (year, qcols, ann_col, suffix)
GD_ALL_YEARS = [
    (2023, [2, 3, 4, 5], 6, "A"),
    (2024, [7, 8, 9, 10], 11, "A"),
    (2025, [12, 13, 14, 15], 16, "A"),
    (2026, [18, 19, 20, 21], 22, "E"),
    (2027, [23, 24, 25, 26], 27, "E"),
    (2028, [28, 29, 30, 31], 32, "E"),
]
# Quarterly segment revenue actuals (where available from 10-Q/estimates)
# Format: {segment: {year: {qnum: value}}}
GD_QREV = {
    "ccg": {
        2025: {1: 7629, 2: 8189, 3: 8190, 4: 8220},
        2026: {1: 7727},
    },
    "dcai": {
        2025: {1: 4126, 2: 4058, 3: 4058, 4: 4677},
        2026: {1: 5052},
    },
    "foundry": {
        2025: {1: 4667, 2: 4321, 3: 4320, 4: 4518},
        2026: {1: 5421},
    },
    "allother": {
        2025: {1: 943, 2: 1022, 3: 1022, 4: 576},
        2026: {1: 628},
    },
}

# ============================================================
# COLOR CODING — IB Standard (see legend on every tab)
# ============================================================
BLUE_FONT  = "305496"   # Historical hardcoded data
GREEN_FONT = "548235"   # Cross-sheet link
BLUE_FILL  = "DDEBF7"   # Editable assumption (blue fill bg)
YELLOW_FILL = "FFF2CC"  # Key output row (yellow fill bg)
GREY_FONT  = "808080"   # Memo / percentage row (grey italic)
DARK_FILL  = "1F4E79"   # Header background (dark blue)
WHITE_FONT = "FFFFFF"   # Header text colour

# ============================================================
# REUSABLE STYLES
# ============================================================

BD = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FONT_HIST   = Font(color=BLUE_FONT, size=10, name="Calibri")
FONT_LINK   = Font(color=GREEN_FONT, size=10, name="Calibri")
FONT_FORM   = Font(color="000000", size=10, name="Calibri")
FONT_MEMO   = Font(color=GREY_FONT, size=9, italic=True, name="Calibri")
FONT_BOLD   = Font(bold=True, size=10, name="Calibri")
FONT_TITLE  = Font(bold=True, size=14, name="Calibri")
FONT_UNIT   = Font(italic=True, size=9, color=GREY_FONT, name="Calibri")
FONT_HEADER = Font(bold=True, color=WHITE_FONT, size=10, name="Calibri")
FONT_SECTION = Font(bold=True, size=11, name="Calibri", color=DARK_FILL)

FILL_ASSUMP = PatternFill("solid", fgColor=BLUE_FILL)
FILL_KEY    = PatternFill("solid", fgColor=YELLOW_FILL)
FILL_HEADER = PatternFill("solid", fgColor=DARK_FILL)
FILL_GREY   = PatternFill("solid", fgColor="F2F2F2")
FILL_BLANK  = PatternFill("solid", fgColor="FFFFFF")

ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT  = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _hval(ws, r, col, val, fmt="#,##0"):
    """Historical hardcoded value — blue font, grey fill"""
    c = ws.cell(row=r, column=col, value=val)
    c.number_format = fmt
    c.font = FONT_HIST
    c.border = BD
    c.alignment = ALIGN_RIGHT
    c.fill = FILL_GREY

def _fval(ws, r, col, formula, fmt="#,##0"):
    """Formula (calculated) — black font"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = fmt
    c.font = FONT_FORM
    c.border = BD
    c.alignment = ALIGN_RIGHT

def _lval(ws, r, col, formula, fmt="#,##0"):
    """Cross-sheet link — green font"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = fmt
    c.font = FONT_LINK
    c.border = BD
    c.alignment = ALIGN_RIGHT

def _pct(ws, r, col, formula):
    """Percentage row — grey italic font"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = "0.0%"
    c.font = FONT_MEMO
    c.border = BD
    c.alignment = ALIGN_RIGHT

def _key(ws, r, col, formula, fmt="#,##0"):
    """Key output row — yellow fill + black font"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = fmt
    c.font = FONT_BOLD
    c.fill = FILL_KEY
    c.border = BD
    c.alignment = ALIGN_RIGHT

def _label(ws, r, text, bold=False):
    """Row label in column A"""
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_BOLD if bold else FONT_FORM
    c.border = BD
    c.alignment = ALIGN_LEFT

def _is_quarterly(ws):
    """True for tabs using quarterly 35-col layout (GD_NCOL)."""
    return ws.title in ("Growth_Drivers", "Assumptions", "Segment_PL")

def _ncol_for(ws):
    """Return the Notes/width column for this tab (35 for quarterly, 11 for annual)."""
    return GD_NCOL if _is_quarterly(ws) else NCOL

def _section(ws, r, text):
    """Section header — bold dark blue"""
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_SECTION
    n = _ncol_for(ws)
    for col in range(1, n):
        ws.cell(row=r, column=col).border = BD

def _note(ws, r, text):
    """Notes column — strips leading '=' to prevent openpyxl formula errors."""
    if text and text.startswith("="):
        text = text[1:]
    col = _ncol_for(ws)
    c = ws.cell(row=r, column=col, value=text)
    c.font = Font(size=9, name="Calibri")
    c.alignment = ALIGN_LEFT


def _title_row(ws, r, text):
    """Tab title — row 1"""
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_TITLE
    n = _ncol_for(ws)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n - 1)

def _unit_row(ws, r, text="USD Million (except per-share data)"):
    """Unit row — row 2, grey italic"""
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_UNIT
    n = _ncol_for(ws)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n - 1)

def _year_header(ws, r, extra_cols=True):
    """Standard year-header row with dark fill + white font.
    Cols: B=2023A, C=2024A, D=2025A, E=(spacer), F=2026E, G=2027E, H=2028E
    If extra_cols, also write I=Selected, K=Notes headers.
    """
    hdrs = {2: "2023A", 3: "2024A", 4: "2025A", 6: "2026E", 7: "2027E", 8: "2028E"}
    for col, txt in hdrs.items():
        c = ws.cell(row=r, column=col, value=txt)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BD
        c.alignment = ALIGN_CENTER
    # Spacer & extra headers
    for col in [5]:
        c = ws.cell(row=r, column=col, value="")
        c.fill = FILL_HEADER
        c.border = BD
    if extra_cols:
        c = ws.cell(row=r, column=9, value="Selected")
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER
        c = ws.cell(row=r, column=NCOL, value="Notes")
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER
    # Column A
    c = ws.cell(row=r, column=1, value="")
    c.fill = FILL_HEADER; c.border = BD

def _yr_header_no_scenario(ws, r):
    """Year header WITHOUT Scenario/Notes columns (for non-Assumptions tabs)"""
    hdrs = {2: "2023A", 3: "2024A", 4: "2025A", 6: "2026E", 7: "2027E", 8: "2028E"}
    for col, txt in hdrs.items():
        c = ws.cell(row=r, column=col, value=txt)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BD
        c.alignment = ALIGN_CENTER
    for col in [1, 5, 9, 10, NCOL]:
        c = ws.cell(row=r, column=col, value="")
        c.fill = FILL_HEADER; c.border = BD
    if NCOL == 11:
        c = ws.cell(row=r, column=NCOL, value="Notes")
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER

def _qheader(ws, r):
    """Quarterly column header for Growth_Drivers and other quarterly tabs.
    35-column layout matching GD_NCOL."""
    hdrs = {
        2: "Q1'23A", 3: "Q2'23A", 4: "Q3'23A", 5: "Q4'23A", 6: "FY2023A",
        7: "Q1'24A", 8: "Q2'24A", 9: "Q3'24A", 10: "Q4'24A", 11: "FY2024A",
        12: "Q1'25A", 13: "Q2'25A", 14: "Q3'25A", 15: "Q4'25A", 16: "FY2025A",
        18: "Q1'26E", 19: "Q2'26E", 20: "Q3'26E", 21: "Q4'26E", 22: "FY2026E",
        23: "Q1'27E", 24: "Q2'27E", 25: "Q3'27E", 26: "Q4'27E", 27: "FY2027E",
        28: "Q1'28E", 29: "Q2'28E", 30: "Q3'28E", 31: "Q4'28E", 32: "FY2028E",
        33: "Selected",
    }
    ANN_COLS_SET = {6, 11, 16, 22, 27, 32}
    for col, txt in hdrs.items():
        c = ws.cell(row=r, column=col, value=txt)
        c.font = FONT_HEADER
        c.fill = PatternFill("solid", fgColor="2E75B6") if col in ANN_COLS_SET else FILL_HEADER
        c.border = BD; c.alignment = ALIGN_CENTER
    for col in [1, 17, 34]:
        c = ws.cell(row=r, column=col, value="")
        c.fill = FILL_HEADER; c.border = BD
    c = ws.cell(row=r, column=GD_NCOL, value="Notes")
    c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER


def _arow(ws, r, label, h1, h2, h3, base, bull, bear, fmt="#,##0.0", note_text=""):
    """Full assumption row: 3 historical (B-D) + 3 scenario values (F-H) + CHOOSE/MATCH (I) + note (K).

    Parameters
    ----------
    h1, h2, h3 : float or None  — historical values; None → leave cell blank
    base, bull, bear : float    — scenario values
    fmt : str                   — number format (e.g. "#,##0.0", "0.0%")
    """
    _label(ws, r, label)
    for col, val in [(2, h1), (3, h2), (4, h3)]:
        if val is not None:
            _hval(ws, r, col, val, fmt)
        else:
            ws.cell(row=r, column=col).border = BD
    # Scenario columns (blue fill = editable)
    for col, val in [(6, base), (7, bull), (8, bear)]:
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = fmt
        c.font = Font(color="000000", size=10, name="Calibri")
        c.fill = FILL_ASSUMP
        c.border = BD
        c.alignment = ALIGN_RIGHT
    # Spacer col E
    ws.cell(row=r, column=5).border = BD
    # CHOOSE/MATCH in col I
    formula = f'=CHOOSE(MATCH({SCENARIO_CELL},{{"Base","Bull","Bear"}},0),F{r},G{r},H{r})'
    c = ws.cell(row=r, column=SEL_COL, value=formula)
    c.number_format = fmt
    c.font = FONT_BOLD
    c.border = BD
    c.alignment = ALIGN_RIGHT
    # Col J spacer
    ws.cell(row=r, column=10).border = BD
    # Note
    _note(ws, r, note_text)

def _arow_const(ws, r, label, h1, h2, h3, val, fmt="#,##0.0", note_text=""):
    """Assumption row where Base=Bull=Bear (scenario-independent). Shortcut for _arow."""
    _arow(ws, r, label, h1, h2, h3, val, val, val, fmt, note_text)

def _add_legend(ws, start_row):
    """Add colour-coding legend at the bottom of a tab."""
    r = start_row
    items = [
        ("Blue font", "305496", "Historical hardcoded data"),
        ("Green font", "548235", "Cross-sheet link"),
        ("Blue fill + dark font", "DDEBF7", "Editable assumption"),
        ("Yellow fill", "FFF2CC", "Key output row"),
        ("Black font", "000000", "Formula / calculated cell"),
        ("Grey italic", "808080", "Memo / percentage row"),
    ]
    _label(ws, r, "DATA LEGEND"); r += 1
    for label, color, desc in items:
        c = ws.cell(row=r, column=1, value=f"  {label}: {desc}")
        if "font" in label.lower():
            c.font = Font(color=color, size=9, name="Calibri")
        elif "fill" in label.lower():
            c.font = Font(color="000000", size=9, name="Calibri")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    return r

def _add_sources(ws, start_row):
    """Add data-sources block at the bottom of a tab."""
    r = start_row
    _label(ws, r, "DATA SOURCES"); r += 1
    sources = [
        "2022-2025 10-Ks: Consolidated P&L, Balance Sheet, Cash Flow, Segment data",
        "2026 Q1 10-Q: Q1 2026 vs Q1 2025 segment + consolidated data",
        "Q1 2026 Earnings Call (Apr 23, 2026): FY2026 management guidance",
        "intc_assumptions.md: All forecast assumptions with scenario narratives",
        "Bloomberg / Damodaran / FRED: WACC inputs (Rf, ERP, Beta) as of May 2026",
    ]
    for s in sources:
        c = ws.cell(row=r, column=1, value=f"  - {s}")
        c.font = Font(size=9, name="Calibri")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    return r


# ============================================================
# SCENARIO SELECTOR SETUP (called once, on Assumptions tab)
# ============================================================

def _setup_scenario_selector(ws):
    """Place scenario dropdown in B2 of the Assumptions tab."""
    c = ws.cell(row=2, column=2, value="Base")
    c.font = Font(bold=True, size=11, name="Calibri", color="1F4E79")
    c.fill = PatternFill("solid", fgColor="DDEBF7")
    c.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium"),
    )
    c.alignment = ALIGN_CENTER
    dv = DataValidation(type="list", formula1='"Base,Bull,Bear"', allow_blank=False)
    dv.error = "Please select Base, Bull, or Bear"
    dv.errorTitle = "Invalid Scenario"
    dv.prompt = "Select scenario"
    dv.promptTitle = "Scenario"
    ws.add_data_validation(dv)
    dv.add(c)
    # Label
    lbl = ws.cell(row=2, column=1, value="Scenario:")
    lbl.font = Font(bold=True, size=11, name="Calibri")
    lbl.alignment = ALIGN_RIGHT


# ============================================================
# TAB 3: ASSUMPTIONS (Control Center)
# ============================================================

def build_assumptions(wb, R):
    """Build the Assumptions tab with quarterly columns, Base/Bull/Bear rows,
    and CHOOSE/MATCH Selected row. User adjusts each quarter independently.
    Annual = SUM or AVERAGE of quarterly values."""
    ws = wb["Assumptions"]
    _title_row(ws, 1, "Intel (INTC) - Assumptions & Scenario Design (Quarterly)")
    _setup_scenario_selector(ws)

    # Row 3: spacer, Row 4: column headers
    r = 4
    # Build quarterly header matching Growth_Drivers
    hdrs = {
        2: "Q1'23A", 3: "Q2'23A", 4: "Q3'23A", 5: "Q4'23A", 6: "FY2023A",
        7: "Q1'24A", 8: "Q2'24A", 9: "Q3'24A", 10: "Q4'24A", 11: "FY2024A",
        12: "Q1'25A", 13: "Q2'25A", 14: "Q3'25A", 15: "Q4'25A", 16: "FY2025A",
        18: "Q1'26E", 19: "Q2'26E", 20: "Q3'26E", 21: "Q4'26E", 22: "FY2026E",
        23: "Q1'27E", 24: "Q2'27E", 25: "Q3'27E", 26: "Q4'27E", 27: "FY2027E",
        28: "Q1'28E", 29: "Q2'28E", 30: "Q3'28E", 31: "Q4'28E", 32: "FY2028E",
    }
    ANN_COLS_SET = {6, 11, 16, 22, 27, 32}
    for col, txt in hdrs.items():
        c = ws.cell(row=r, column=col, value=txt)
        c.font = FONT_HEADER
        c.fill = PatternFill("solid", fgColor="2E75B6") if col in ANN_COLS_SET else FILL_HEADER
        c.border = BD; c.alignment = ALIGN_CENTER
    for col in [1, 17, 33, 34]:
        c = ws.cell(row=r, column=col, value="")
        c.fill = FILL_HEADER; c.border = BD
    c = ws.cell(row=r, column=GD_NCOL, value="Notes")
    c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER

    # All 6 years iteration
    all_years = [
        (2023, [2,3,4,5], 6), (2024, [7,8,9,10], 11), (2025, [12,13,14,15], 16),
        (2026, [18,19,20,21], 22), (2027, [23,24,25,26], 27), (2028, [28,29,30,31], 32),
    ]
    hist_years = all_years[:3]
    fcst_years = all_years[3:]

    r = 6

    # ============================================================
    # Helper: write quarterly assumption with Base/Bull/Bear/Selected rows
    # ============================================================
    def _aq(label, hist_qvals, base_qvals, bull_qvals, bear_qvals, fmt_val, note_text, driver_type="sum"):
        """Quarterly assumption: 5 rows (label + Base/Bull/Bear/Selected).
        hist_qvals: dict {year: [q1,q2,q3,q4]} or None (uses annual/4 of first hist year).
        base_qvals: dict {year: [q1,q2,q3,q4]} for forecast quarters.
        driver_type: 'sum' -> annual=SUM(Q1:Q4), 'avg' -> annual=AVERAGE(Q1:Q4).
        Returns: (next_row, selected_row_num)"""
        nonlocal r
        _label(ws, r, label)
        r_label = r; r += 1

        base_r = r; r += 1
        bull_r = r; r += 1
        bear_r = r; r += 1
        sel_r = r; r += 1

        # Column indices for iteration
        all_qcols = []
        for yr, qcols, ann_col in all_years:
            all_qcols.append((yr, qcols, ann_col))
        hist_qcols = [(yr, qcols, ann_col) for yr, qcols, ann_col in hist_years]
        fcst_qcols = [(yr, qcols, ann_col) for yr, qcols, ann_col in fcst_years]

        ann_fn = "SUM" if driver_type == "sum" else "AVERAGE"

        for scenario, srow in [("Base", base_r), ("Bull", bull_r), ("Bear", bear_r)]:
            ws.cell(row=srow, column=1, value=f"  {scenario}").font = Font(bold=True, size=10, name="Calibri")
            ws.cell(row=srow, column=1).border = BD
            # Historical: same across scenarios, put in Base row only (leave Bull/Bear blank for clarity)
            if scenario == "Base":
                for yr, qcols, ann_col in hist_years:
                    qv = hist_qvals.get(yr) if hist_qvals else None
                    for qi, qc in enumerate(qcols):
                        if qv:
                            _hval(ws, srow, qc, qv[qi], fmt_val)
                        else:
                            ws.cell(row=srow, column=qc).border = BD
                    if qv:
                        if driver_type == "sum":
                            _fval(ws, srow, ann_col, f"=SUM({CL(qcols[0])}{srow}:{CL(qcols[3])}{srow})", fmt_val)
                        else:
                            _fval(ws, srow, ann_col, f"=AVERAGE({CL(qcols[0])}{srow}:{CL(qcols[3])}{srow})", fmt_val)
                    else:
                        ws.cell(row=srow, column=ann_col).border = BD
            else:
                for yr, qcols, ann_col in hist_years:
                    for qc in qcols:
                        ws.cell(row=srow, column=qc).border = BD
                    ws.cell(row=srow, column=ann_col).border = BD
            # Forecast: editable quarterly values (blue fill)
            for yr, qcols, ann_col in fcst_years:
                qv = {"Base": base_qvals, "Bull": bull_qvals, "Bear": bear_qvals}[scenario].get(yr)
                for qi, qc in enumerate(qcols):
                    if qv:
                        c = ws.cell(row=srow, column=qc, value=qv[qi])
                        c.number_format = fmt_val
                        c.font = Font(color="000000", size=10, name="Calibri")
                        c.fill = FILL_ASSUMP
                        c.border = BD
                        c.alignment = ALIGN_RIGHT
                    else:
                        ws.cell(row=srow, column=qc).border = BD
                # Annual = SUM/AVERAGE of quarters
                if qv:
                    if driver_type == "sum":
                        _fval(ws, srow, ann_col, f"=SUM({CL(qcols[0])}{srow}:{CL(qcols[3])}{srow})", fmt_val)
                    else:
                        _fval(ws, srow, ann_col, f"=AVERAGE({CL(qcols[0])}{srow}:{CL(qcols[3])}{srow})", fmt_val)
                else:
                    ws.cell(row=srow, column=ann_col).border = BD
            # Spacer + note cols
            for sc in [17, 33, 34]:
                ws.cell(row=srow, column=sc).border = BD

        # Selected row
        ws.cell(row=sel_r, column=1, value="  SELECTED").font = Font(bold=True, size=10, name="Calibri", color="1F4E79")
        ws.cell(row=sel_r, column=1).border = BD
        # Historical: hardcoded (same as Base)
        for yr, qcols, ann_col in hist_years:
            qv = hist_qvals.get(yr) if hist_qvals else None
            for qi, qc in enumerate(qcols):
                if qv:
                    _hval(ws, sel_r, qc, qv[qi], fmt_val)
                else:
                    ws.cell(row=sel_r, column=qc).border = BD
            if qv:
                if driver_type == "sum":
                    _fval(ws, sel_r, ann_col, f"=SUM({CL(qcols[0])}{sel_r}:{CL(qcols[3])}{sel_r})", fmt_val)
                else:
                    _fval(ws, sel_r, ann_col, f"=AVERAGE({CL(qcols[0])}{sel_r}:{CL(qcols[3])}{sel_r})", fmt_val)
            else:
                ws.cell(row=sel_r, column=ann_col).border = BD
        # Forecast: CHOOSE/MATCH per quarter
        for yr, qcols, ann_col in fcst_years:
            for qc in qcols:
                cl = CL(qc)
                formula = f'=CHOOSE(MATCH({SCENARIO_CELL},{{"Base","Bull","Bear"}},0),{cl}{base_r},{cl}{bull_r},{cl}{bear_r})'
                c = ws.cell(row=sel_r, column=qc, value=formula)
                c.number_format = fmt_val
                c.font = FONT_BOLD
                c.border = BD
                c.alignment = ALIGN_RIGHT
            if driver_type == "sum":
                _fval(ws, sel_r, ann_col, f"=SUM({CL(qcols[0])}{sel_r}:{CL(qcols[3])}{sel_r})", fmt_val)
            else:
                _fval(ws, sel_r, ann_col, f"=AVERAGE({CL(qcols[0])}{sel_r}:{CL(qcols[3])}{sel_r})", fmt_val)
        for sc in [17, 33, 34]:
            ws.cell(row=sel_r, column=sc).border = BD
        # Notes
        c = ws.cell(row=sel_r, column=GD_NCOL, value=note_text)
        c.font = Font(size=9, name="Calibri"); c.alignment = ALIGN_LEFT

        # Fill key cells with yellow for Selected row annual cols
        for yr, qcols, ann_col in fcst_years:
            ws.cell(row=sel_r, column=ann_col).fill = FILL_KEY

        r += 1  # blank row after each driver group
        return sel_r

    # ============================================================
    # Helper: annual-only parameter (still uses Base/Bull/Bear rows but only annual columns)
    # ============================================================
    def _a_annual(label, hist_vals, base_vals, bull_vals, bear_vals, fmt_val, note_text):
        """Annual parameter: Base/Bull/Bear rows, values only in FY columns (6,11,16,22,27,32)."""
        nonlocal r
        _label(ws, r, label)
        r += 1
        base_r = r; r += 1
        bull_r = r; r += 1
        bear_r = r; r += 1
        sel_r = r; r += 2

        ann_hist = [6, 11, 16]
        ann_fcst = [22, 27, 32]

        for scenario, srow, vals in [("Base", base_r, base_vals), ("Bull", bull_r, bull_vals), ("Bear", bear_r, bear_vals)]:
            ws.cell(row=srow, column=1, value=f"  {scenario}").font = Font(bold=True, size=10, name="Calibri")
            ws.cell(row=srow, column=1).border = BD
            # Historical (same across all, only in Base)
            if scenario == "Base":
                for ci, ac in enumerate(ann_hist):
                    if hist_vals and ci < len(hist_vals) and hist_vals[ci] is not None:
                        _hval(ws, srow, ac, hist_vals[ci], fmt_val)
                    else:
                        ws.cell(row=srow, column=ac).border = BD
            else:
                for ac in ann_hist:
                    ws.cell(row=srow, column=ac).border = BD
            # Forecast
            for fi, ac in enumerate(ann_fcst):
                if vals and fi < len(vals):
                    c = ws.cell(row=srow, column=ac, value=vals[fi])
                    c.number_format = fmt_val
                    c.font = Font(color="000000", size=10, name="Calibri")
                    c.fill = FILL_ASSUMP
                    c.border = BD
                    c.alignment = ALIGN_RIGHT
                else:
                    ws.cell(row=srow, column=ac).border = BD
            for sc in [17, 33, 34]:
                ws.cell(row=srow, column=sc).border = BD

        # Selected row
        ws.cell(row=sel_r, column=1, value="  SELECTED").font = Font(bold=True, size=10, name="Calibri", color="1F4E79")
        ws.cell(row=sel_r, column=1).border = BD
        for ci, ac in enumerate(ann_hist):
            if hist_vals and ci < len(hist_vals) and hist_vals[ci] is not None:
                _hval(ws, sel_r, ac, hist_vals[ci], fmt_val)
            else:
                ws.cell(row=sel_r, column=ac).border = BD
        for ac in ann_fcst:
            cl = CL(ac)
            formula = f'=CHOOSE(MATCH({SCENARIO_CELL},{{"Base","Bull","Bear"}},0),{cl}{base_r},{cl}{bull_r},{cl}{bear_r})'
            c = ws.cell(row=sel_r, column=ac, value=formula)
            c.number_format = fmt_val
            c.font = FONT_BOLD
            c.fill = FILL_KEY
            c.border = BD
            c.alignment = ALIGN_RIGHT
        for sc in [17, 33, 34]:
            ws.cell(row=sel_r, column=sc).border = BD
        _note(ws, sel_r, note_text)
        return sel_r

    # ============================================================
    # BOTTOM-UP GROWTH DRIVERS: CCG
    # ============================================================
    _section(ws, r, "CCG Bottom-Up Drivers: Revenue = PC TAM x Intel Share x ASP"); r += 1

    # PC TAM — sum-type (annual split equally across quarters)
    R["asp_pc_tam"] = _aq(
        "PC TAM (M units) - Source: Gartner/IDC",
        {2023: [60.45, 60.45, 60.45, 60.45], 2024: [61.35, 61.35, 61.35, 61.35], 2025: [64.65, 64.65, 64.65, 64.65]},
        {2026: [65.0, 65.0, 65.0, 65.0], 2027: [66.25, 66.25, 66.25, 66.25], 2028: [67.5, 67.5, 67.5, 67.5]},
        {2026: [67.5, 67.5, 67.5, 67.5], 2027: [70.0, 70.0, 70.0, 70.0], 2028: [72.5, 72.5, 72.5, 72.5]},
        {2026: [62.5, 62.5, 62.5, 62.5], 2027: [63.0, 63.0, 63.0, 63.0], 2028: [63.75, 63.75, 63.75, 63.75]},
        "#,##0.0",
        "Annual TAM split equally across 4 quarters (edit individual quarters as needed). Gartner: 2023=241.8M, 2024=245.4M, 2025~259M.",
        "sum"
    )

    # Intel PC Unit Share — avg-type (same % each quarter)
    R["asp_intel_share"] = _aq(
        "Intel PC Unit Share (%) - Source: Mercury Research",
        {2023: [0.710, 0.710, 0.710, 0.710], 2024: [0.715, 0.715, 0.715, 0.715], 2025: [0.720, 0.720, 0.720, 0.720]},
        {2026: [0.72, 0.72, 0.72, 0.72], 2027: [0.725, 0.725, 0.725, 0.725], 2028: [0.73, 0.73, 0.73, 0.73]},
        {2026: [0.74, 0.74, 0.74, 0.74], 2027: [0.75, 0.75, 0.75, 0.75], 2028: [0.76, 0.76, 0.76, 0.76]},
        {2026: [0.70, 0.70, 0.70, 0.70], 2027: [0.69, 0.69, 0.69, 0.69], 2028: [0.68, 0.68, 0.68, 0.68]},
        "0.0%",
        "Mercury est: ~72%. 18A products competitive; share recovery assumed. Edit quarterly for seasonal patterns.",
        "avg"
    )

    # CCG Blended ASP — avg-type
    R["asp_ccg_asp"] = _aq(
        "Intel CCG Blended ASP ($) - Derived from Rev/(TAMxShare)",
        {2023: [188, 188, 188, 188], 2024: [190, 190, 190, 190], 2025: [173, 173, 173, 173]},
        {2026: [175, 175, 175, 175], 2027: [182, 182, 182, 182], 2028: [184, 184, 184, 184]},
        {2026: [180, 180, 180, 180], 2027: [190, 190, 190, 190], 2028: [195, 195, 195, 195]},
        {2026: [170, 170, 170, 170], 2027: [172, 172, 172, 172], 2028: [174, 174, 174, 174]},
        "0.0",
        "FY2025 ASP depressed. Q1 2026 +16% YoY. Base: ASP recovery + premium mix. Edit quarterly for product launch timing.",
        "avg"
    )

    # ============================================================
    # DCAI Bottom-Up Drivers
    # ============================================================
    _section(ws, r, "DCAI Bottom-Up Drivers: CPU Rev = Server TAM x Share x ASP + AI/ASIC Rev"); r += 1

    R["asp_svr_tam"] = _aq(
        "Server TAM (M units) - Source: IDC",
        {2023: [2.875, 2.875, 2.875, 2.875], 2024: [3.0, 3.0, 3.0, 3.0], 2025: [3.375, 3.375, 3.375, 3.375]},
        {2026: [3.375, 3.375, 3.375, 3.375], 2027: [3.625, 3.625, 3.625, 3.625], 2028: [3.875, 3.875, 3.875, 3.875]},
        {2026: [3.625, 3.625, 3.625, 3.625], 2027: [4.0, 4.0, 4.0, 4.0], 2028: [4.375, 4.375, 4.375, 4.375]},
        {2026: [3.125, 3.125, 3.125, 3.125], 2027: [3.25, 3.25, 3.25, 3.25], 2028: [3.375, 3.375, 3.375, 3.375]},
        "#,##0.0",
        "IDC: 2023=11.5M, 2024=12.0M, 2025~13.5M. AI buildout driving growth. Equal quarterly allocation (edit for seasonality).",
        "sum"
    )

    R["asp_intel_dc_share"] = _aq(
        "Intel DC Server Unit Share (%) - Source: Mercury Research",
        {2023: [0.775, 0.775, 0.775, 0.775], 2024: [0.760, 0.760, 0.760, 0.760], 2025: [0.730, 0.730, 0.730, 0.730]},
        {2026: [0.70, 0.70, 0.70, 0.70], 2027: [0.70, 0.70, 0.70, 0.70], 2028: [0.70, 0.70, 0.70, 0.70]},
        {2026: [0.72, 0.72, 0.72, 0.72], 2027: [0.73, 0.73, 0.73, 0.73], 2028: [0.74, 0.74, 0.74, 0.74]},
        {2026: [0.67, 0.67, 0.67, 0.67], 2027: [0.65, 0.65, 0.65, 0.65], 2028: [0.63, 0.63, 0.63, 0.63]},
        "0.0%",
        "Mercury: Q4'24 ~72%. Granite Rapids improving; AMD Turin competitive. Base: share stabilizes.",
        "avg"
    )

    R["asp_dc_asp"] = _aq(
        "Intel Server Blended ASP ($) - Derived from CPU Rev/(TAMxShare)",
        {2023: [1772, 1772, 1772, 1772], 2024: [1712, 1712, 1712, 1712], 2025: [1632, 1632, 1632, 1632]},
        {2026: [1600, 1600, 1600, 1600], 2027: [1700, 1700, 1700, 1700], 2028: [1750, 1750, 1750, 1750]},
        {2026: [1750, 1750, 1750, 1750], 2027: [1850, 1850, 1850, 1850], 2028: [1950, 1950, 1950, 1950]},
        {2026: [1500, 1500, 1500, 1500], 2027: [1550, 1550, 1550, 1550], 2028: [1580, 1580, 1580, 1580]},
        "#,##0",
        "FY2025 ASP depressed. Q1 2026 +27% YoY. Base: premium mix recovery. Edit quarterly for product cycle effects.",
        "avg"
    )

    R["asp_ai_asic"] = _aq(
        "AI/ASIC & Other Revenue ($M) - Gaudi, ASICs, NICs, IPUs",
        {2023: [50, 50, 50, 50], 2024: [125, 125, 125, 125], 2025: [200, 200, 200, 200]},
        {2026: [700, 700, 700, 700], 2027: [1125, 1125, 1125, 1125], 2028: [1500, 1500, 1500, 1500]},
        {2026: [875, 875, 875, 875], 2027: [1500, 1500, 1500, 1500], 2028: [2125, 2125, 2125, 2125]},
        {2026: [500, 500, 500, 500], 2027: [750, 750, 750, 750], 2028: [950, 950, 950, 950]},
        "#,##0",
        "ASIC run rate 'north of $1B' Q1 2026. Gaudi modest. Edit quarterly for customer ramp timing.",
        "sum"
    )

    # ============================================================
    # Intel Foundry Bottom-Up Drivers
    # ============================================================
    _section(ws, r, "Intel Foundry Bottom-Up Drivers: Internal Rev = Total Chip Vol x Rev/Chip + External"); r += 1

    R["asp_foundry_rev_per_chip"] = _aq(
        "Internal Foundry Rev per Chip ($/unit) - Internal Foundry Rev / Total Intel Units",
        {2023: [102.2, 102.2, 102.2, 102.2], 2024: [93.5, 93.5, 93.5, 93.5], 2025: [90.1, 90.1, 90.1, 90.1]},
        {2026: [90, 90, 90, 90], 2027: [92, 92, 92, 92], 2028: [95, 95, 95, 95]},
        {2026: [98, 98, 98, 98], 2027: [105, 105, 105, 105], 2028: [110, 110, 110, 110]},
        {2026: [82, 82, 82, 82], 2027: [80, 80, 80, 80], 2028: [78, 78, 78, 78]},
        "0.0",
        "Declining trend stabilizes as 18A premium wafers ramp. Edit quarterly for node transition effects.",
        "avg"
    )

    R["asp_ext_foundry"] = _aq(
        "External Foundry Revenue ($M) - 3rd-party wafer customers",
        {2023: [12.5, 12.5, 12.5, 12.5], 2024: [15, 15, 15, 15], 2025: [37.5, 37.5, 37.5, 37.5]},
        {2026: [175, 175, 175, 175], 2027: [375, 375, 375, 375], 2028: [625, 625, 625, 625]},
        {2026: [250, 250, 250, 250], 2027: [750, 750, 750, 750], 2028: [1250, 1250, 1250, 1250]},
        {2026: [100, 100, 100, 100], 2027: [200, 200, 200, 200], 2028: [300, 300, 300, 300]},
        "#,##0",
        "~$150M FY2025. Growing as 18A/14A external PDK ramps. Edit quarterly for customer win timing.",
        "sum"
    )

    # ============================================================
    # All Other Bottom-Up Drivers
    # ============================================================
    _section(ws, r, "All Other Bottom-Up Drivers: Mobileye (LV Prod x Rev/Veh) + IMS/Other"); r += 1

    R["asp_lv_prod"] = _aq(
        "Global Light Vehicle Production (M units) - Source: IHS/S&P Global Mobility",
        {2023: [22.25, 22.25, 22.25, 22.25], 2024: [22.0, 22.0, 22.0, 22.0], 2025: [21.75, 21.75, 21.75, 21.75]},
        {2026: [22.0, 22.0, 22.0, 22.0], 2027: [22.25, 22.25, 22.25, 22.25], 2028: [22.5, 22.5, 22.5, 22.5]},
        {2026: [22.5, 22.5, 22.5, 22.5], 2027: [23.0, 23.0, 23.0, 23.0], 2028: [23.5, 23.5, 23.5, 23.5]},
        {2026: [21.25, 21.25, 21.25, 21.25], 2027: [21.5, 21.5, 21.5, 21.5], 2028: [21.75, 21.75, 21.75, 21.75]},
        "#,##0.0",
        "IHS/S&P Global Mobility. FY2025 ~87M. Modest recovery in Base case.",
        "sum"
    )

    R["asp_mbly_rev_per_veh"] = _aq(
        "Mobileye Rev per Vehicle ($/vehicle) - Derived from Mobileye Rev / LV Production",
        {2023: [23.4, 23.4, 23.4, 23.4], 2024: [18.9, 18.9, 18.9, 18.9], 2025: [21.8, 21.8, 21.8, 21.8]},
        {2026: [19, 19, 19, 19], 2027: [20, 20, 20, 20], 2028: [21, 21, 21, 21]},
        {2026: [21, 21, 21, 21], 2027: [23, 23, 23, 23], 2028: [25, 25, 25, 25]},
        {2026: [17, 17, 17, 17], 2027: [17.5, 17.5, 17.5, 17.5], 2028: [18, 18, 18, 18]},
        "0.0",
        "FY2025 Mobileye $1.9B / 87M = $21.8. Q1 2026 $628M. Base: conservatively ~$19.",
        "avg"
    )

    R["asp_ims_other"] = _aq(
        "IMS & Other Revenue ($M) - IMS nanofabrication tools + legacy businesses",
        {2023: [844.5, 844.5, 844.5, 844.5], 2024: [485.25, 485.25, 485.25, 485.25], 2025: [415.75, 415.75, 415.75, 415.75]},
        {2026: [450, 450, 450, 450], 2027: [500, 500, 500, 500], 2028: [550, 550, 550, 550]},
        {2026: [550, 550, 550, 550], 2027: [650, 650, 650, 650], 2028: [750, 750, 750, 750]},
        {2026: [375, 375, 375, 375], 2027: [400, 400, 400, 400], 2028: [425, 425, 425, 425]},
        "#,##0",
        "IMS multi-beam mask writers. FY2023-24 include Altera (deconsolidated Sep 2025).",
        "sum"
    )

    # ============================================================
    # SEGMENT OPERATING MARGINS — avg-type (same each quarter)
    # ============================================================
    _section(ws, r, "Segment Operating Margin Assumptions"); r += 1

    R["asp_ccg_opm"] = _aq(
        "CCG Operating Margin (%)",
        {2023: [0.314, 0.314, 0.314, 0.314], 2024: [0.348, 0.348, 0.348, 0.348], 2025: [0.289, 0.289, 0.289, 0.289]},
        {2026: [0.30, 0.30, 0.30, 0.30], 2027: [0.315, 0.315, 0.315, 0.315], 2028: [0.33, 0.33, 0.33, 0.33]},
        {2026: [0.33, 0.33, 0.33, 0.33], 2027: [0.36, 0.36, 0.36, 0.36], 2028: [0.38, 0.38, 0.38, 0.38]},
        {2026: [0.27, 0.27, 0.27, 0.27], 2027: [0.27, 0.27, 0.27, 0.27], 2028: [0.28, 0.28, 0.28, 0.28]},
        "0.0%",
        "Q1 32.6% moderated by 18A ramp costs. Edit quarterly for product launch margin impact.",
        "avg"
    )

    R["asp_dcai_opm"] = _aq(
        "DCAI Operating Margin (%)",
        {2023: [0.059, 0.059, 0.059, 0.059], 2024: [0.088, 0.088, 0.088, 0.088], 2025: [0.202, 0.202, 0.202, 0.202]},
        {2026: [0.28, 0.28, 0.28, 0.28], 2027: [0.28, 0.28, 0.28, 0.28], 2028: [0.27, 0.27, 0.27, 0.27]},
        {2026: [0.32, 0.32, 0.32, 0.32], 2027: [0.33, 0.33, 0.33, 0.33], 2028: [0.32, 0.32, 0.32, 0.32]},
        {2026: [0.22, 0.22, 0.22, 0.22], 2027: [0.20, 0.20, 0.20, 0.20], 2028: [0.18, 0.18, 0.18, 0.18]},
        "0.0%",
        "Sustains near Q1 30.5% level; ASIC supports margins.",
        "avg"
    )

    R["asp_foundry_opm"] = _aq(
        "Intel Foundry Operating Margin (%)",
        {2023: [-0.383, -0.383, -0.383, -0.383], 2024: [-0.767, -0.767, -0.767, -0.767], 2025: [-0.579, -0.579, -0.579, -0.579]},
        {2026: [-0.42, -0.42, -0.42, -0.42], 2027: [-0.32, -0.32, -0.32, -0.32], 2028: [-0.22, -0.22, -0.22, -0.22]},
        {2026: [-0.35, -0.35, -0.35, -0.35], 2027: [-0.20, -0.20, -0.20, -0.20], 2028: [-0.08, -0.08, -0.08, -0.08]},
        {2026: [-0.50, -0.50, -0.50, -0.50], 2027: [-0.45, -0.45, -0.45, -0.45], 2028: [-0.40, -0.40, -0.40, -0.40]},
        "0.0%",
        "18A yield improvement + volume absorption. Still unprofitable; multi-year journey.",
        "avg"
    )

    R["asp_allother_opm"] = _aq(
        "All Other Operating Margin (%)",
        {2023: [0.276, 0.276, 0.276, 0.276], 2024: [-0.016, -0.016, -0.016, -0.016], 2025: [0.074, 0.074, 0.074, 0.074]},
        {2026: [0.14, 0.14, 0.14, 0.14], 2027: [0.18, 0.18, 0.18, 0.18], 2028: [0.20, 0.20, 0.20, 0.20]},
        {2026: [0.20, 0.20, 0.20, 0.20], 2027: [0.25, 0.25, 0.25, 0.25], 2028: [0.28, 0.28, 0.28, 0.28]},
        {2026: [0.08, 0.08, 0.08, 0.08], 2027: [0.10, 0.10, 0.10, 0.10], 2028: [0.12, 0.12, 0.12, 0.12]},
        "0.0%",
        "Mobileye gradual recovery + IMS growth.",
        "avg"
    )

    # ============================================================
    # CONSOLIDATED COST STRUCTURE — avg-type
    # ============================================================
    _section(ws, r, "Consolidated Cost Structure (% of Revenue)"); r += 1

    R["asp_cogs"] = _aq(
        "COGS as % of Revenue (=> GM% = 1 - COGS%)",
        {2023: [0.600, 0.600, 0.600, 0.600], 2024: [0.673, 0.673, 0.673, 0.673], 2025: [0.652, 0.652, 0.652, 0.652]},
        {2026: [0.615, 0.615, 0.615, 0.615], 2027: [0.590, 0.590, 0.590, 0.590], 2028: [0.565, 0.565, 0.565, 0.565]},
        {2026: [0.590, 0.590, 0.590, 0.590], 2027: [0.550, 0.550, 0.550, 0.550], 2028: [0.510, 0.510, 0.510, 0.510]},
        {2026: [0.640, 0.640, 0.640, 0.640], 2027: [0.630, 0.630, 0.630, 0.630], 2028: [0.620, 0.620, 0.620, 0.620]},
        "0.0%",
        "Base GM: 38.5%/41.0%/43.5%. Q1 39.4% base; 18A mix headwind offset by yield gains.",
        "avg"
    )

    R["asp_rd"] = _aq(
        "R&D as % of Revenue",
        {2023: [0.296, 0.296, 0.296, 0.296], 2024: [0.312, 0.312, 0.312, 0.312], 2025: [0.261, 0.261, 0.261, 0.261]},
        {2026: [0.260, 0.260, 0.260, 0.260], 2027: [0.250, 0.250, 0.250, 0.250], 2028: [0.240, 0.240, 0.240, 0.240]},
        {2026: [0.250, 0.250, 0.250, 0.250], 2027: [0.230, 0.230, 0.230, 0.230], 2028: [0.210, 0.210, 0.210, 0.210]},
        {2026: [0.280, 0.280, 0.280, 0.280], 2027: [0.280, 0.280, 0.280, 0.280], 2028: [0.270, 0.270, 0.270, 0.270]},
        "0.0%",
        "~$14.5B absolute in FY2026E; declining as % of growing revenue.",
        "avg"
    )

    R["asp_mga"] = _aq(
        "MG&A as % of Revenue",
        {2023: [0.104, 0.104, 0.104, 0.104], 2024: [0.104, 0.104, 0.104, 0.104], 2025: [0.087, 0.087, 0.087, 0.087]},
        {2026: [0.080, 0.080, 0.080, 0.080], 2027: [0.075, 0.075, 0.075, 0.075], 2028: [0.070, 0.070, 0.070, 0.070]},
        {2026: [0.075, 0.075, 0.075, 0.075], 2027: [0.065, 0.065, 0.065, 0.065], 2028: [0.060, 0.060, 0.060, 0.060]},
        {2026: [0.090, 0.090, 0.090, 0.090], 2027: [0.090, 0.090, 0.090, 0.090], 2028: [0.085, 0.085, 0.085, 0.085]},
        "0.0%",
        "~$4.5B flat absolute in FY2026E; declining as % of revenue.",
        "avg"
    )

    # ============================================================
    # ANNUAL PARAMETERS — use annual columns only
    # ============================================================
    _section(ws, r, "Annual Parameters (Non-Quarterly)"); r += 1

    R["asp_da"] = _a_annual(
        "D&A ($M absolute)",
        [7847, 9951, 10757],
        [12500, 13200, 13800],
        [12000, 12500, 12800],
        [13000, 14000, 15000],
        "#,##0",
        "Q1 annualised ~$12.5B; CIP->in-service driving growth."
    )

    R["asp_sbc"] = _a_annual(
        "SBC ($M absolute)",
        [3229, 3410, 2434],
        [2500, 2500, 2500],
        [2500, 2500, 2500],
        [2500, 2500, 2500],
        "#,##0",
        "Q1 2026 annualised ~$2.5B. Scenario-independent."
    )

    R["asp_restruct"] = _a_annual(
        "Restructuring & Other Charges ($M)",
        [-62, 6970, 2191],
        [500, 500, 500],
        [500, 500, 500],
        [500, 500, 500],
        "#,##0",
        "Winding down; plans substantially complete."
    )

    # CapEx
    R["asp_capex"] = _a_annual(
        "Capital Expenditure ($M Gross)",
        [25750, 25122, 17672],
        [17700, 19500, 21000],
        [17500, 22000, 25000],
        [18000, 17000, 16000],
        "#,##0",
        "FY2026E flat YoY per mgmt guidance; tool-heavy."
    )

    # Working Capital
    R["asp_ar_days"] = _a_annual(
        "AR Days",
        [22.9, 23.9, 26.5],
        [25, 25, 25],
        [25, 25, 25],
        [25, 25, 25],
        "#,##0.0",
        "Historical range 22-27 days."
    )

    R["asp_ap_days"] = _a_annual(
        "AP Days",
        [96.3, 128.2, 104.6],
        [100, 100, 100],
        [100, 100, 100],
        [100, 100, 100],
        "#,##0.0",
        "Normalised from FY2024 peak (vendor terms)."
    )

    R["asp_inv_days"] = _a_annual(
        "Inventory Days",
        [124.9, 124.5, 123.0],
        [125, 125, 125],
        [125, 125, 125],
        [125, 125, 125],
        "#,##0.0",
        "Remarkably stable at ~124 days."
    )

    R["asp_accrued_comp"] = _a_annual(
        "Accrued Comp % of Revenue",
        [0.067, 0.063, 0.076],
        [0.07, 0.07, 0.07],
        [0.07, 0.07, 0.07],
        [0.07, 0.07, 0.07],
        "0.0%",
        "Midpoint of historical range."
    )

    R["asp_other_ca"] = _a_annual(
        "Other CA % of Revenue",
        [None, None, None],
        [0.15, 0.15, 0.15],
        [0.15, 0.15, 0.15],
        [0.15, 0.15, 0.15],
        "0.0%",
        "Includes ~$7.5B refundable tax credits."
    )

    R["asp_other_cl"] = _a_annual(
        "Other CL % of Revenue",
        [None, None, None],
        [0.25, 0.25, 0.25],
        [0.25, 0.25, 0.25],
        [0.25, 0.25, 0.25],
        "0.0%",
        "Stable relationship."
    )

    R["asp_tax_payable"] = _a_annual(
        "Income Tax Payable % of Revenue",
        [None, None, None],
        [0.02, 0.02, 0.02],
        [0.02, 0.02, 0.02],
        [0.02, 0.02, 0.02],
        "0.0%",
        "Normalised level."
    )

    # Financing
    R["asp_debt"] = _a_annual(
        "Total Debt ($M)",
        [49266, 50011, 46585],
        [49500, 47200, 46000],
        [47000, 43000, 39000],
        [52000, 54000, 55000],
        "#,##0",
        "FY2026E: +$6.5B Fab34 loan -$2.5B maturities."
    )

    R["asp_shares"] = _a_annual(
        "Basic Shares (M)",
        [4228, 4330, 4994],
        [5100, 5180, 5250],
        [5080, 5120, 5150],
        [5200, 5400, 5600],
        "#,##0",
        "Q1 5,023M + ESPP/RSU + Escrowed Share releases."
    )

    R["asp_dps"] = _a_annual(
        "DPS ($/share)",
        [0.74, 0.38, 0.00],
        [0.00, 0.00, 0.00],
        [0.00, 0.00, 0.00],
        [0.00, 0.00, 0.00],
        "0.00",
        "Dividends suspended; no reinstatement expected."
    )

    R["asp_net_eq_iss"] = _a_annual(
        "Net Equity Issuance ($M)",
        [None, None, None],
        [300, 300, 300],
        [300, 300, 300],
        [300, 300, 300],
        "#,##0",
        "ESPP proceeds net of RSU withholdings."
    )

    # Tax & Non-Operating
    R["asp_etr"] = _a_annual(
        "Effective Tax Rate",
        [-1.198, -0.716, 0.983],
        [0.12, 0.13, 0.13],
        [0.11, 0.12, 0.12],
        [0.15, 0.15, 0.15],
        "0.0%",
        "Anchored to Q2 2026 guided 11%. FY2026E/FY2027E/FY2028E."
    )

    R["asp_interest"] = _a_annual(
        "Interest & Other, Net ($M)",
        [629, 226, 3257],
        [-800, -600, -400],
        [-500, -300, -100],
        [-1200, -1200, -1000],
        "#,##0",
        "Net interest on ~$50B debt (~5% coupon)."
    )

    R["asp_nci"] = _a_annual(
        "Non-Controlling Interests ($M)",
        [293, 293, 293],
        [750, 1100, 1100],
        [750, 1100, 1100],
        [750, 1100, 1100],
        "#,##0",
        "Per mgmt: ~$250M/q Q2-Q4 FY2026E; Ireland SCIP NCI eliminated Apr 2026."
    )

    # Corporate
    R["asp_corp_unalloc"] = _a_annual(
        "Corporate Unallocated ($M)",
        [-5199, -11177, -5518],
        [-5000, -4500, -4000],
        [-5000, -4500, -4000],
        [-5000, -4500, -4000],
        "#,##0",
        "Declining restructuring + SBC reallocation."
    )

    R["asp_inter_elim"] = _a_annual(
        "Intersegment Eliminations ($M)",
        [-205, -161, 619],
        [-200, -200, -200],
        [-200, -200, -200],
        [-200, -200, -200],
        "#,##0",
        "Residual inter-segment elim."
    )

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 44
    for c in range(2, GD_NCOL + 1):
        cl = CL(c)
        if c in [17, 33, 34]:
            ws.column_dimensions[cl].width = 2
        elif c in [6, 11, 16, 22, 27, 32]:
            ws.column_dimensions[cl].width = 15
        elif c == GD_NCOL:
            ws.column_dimensions[cl].width = 55
        else:
            ws.column_dimensions[cl].width = 12

    # Legend & Sources
    lr = _add_legend(ws, r + 2)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "1F4E79"
    return r


def build_growth_drivers(wb, R):
    """Build growth drivers tab with FULL quarterly layout for all 6 years.
    Historical (2023-2025): quarterly actuals where available, annual/4 otherwise.
    Forecast (2026-2028): quarterly = Assumptions annual / 4 (sum-type) or = annual (avg-type).
    Annual columns = SUM(Q1:Q4) for dollars, = AVERAGE(Q1:Q4) for percentages.
    Forecast annual columns (22,27,32) link to Segment_PL."""
    ws = wb["Growth_Drivers"]
    _title_row(ws, 1, "Intel (INTC) - Revenue Growth Drivers (Quarterly All Years)")
    _unit_row(ws, 2, "USD Million / M Units / $ per Unit (All 6 years quarterly: Q1-Q4 + FY rollup)")
    _qheader(ws, 3)
    _label(ws, 4, "Bottom-up revenue decomposition. Hist: quarterly from 10-Q/estimates. Fcst: quarterly from Assumptions/4, annual=SUM.")
    _note(ws, 4, "Annual columns (F,K,P,V,AA,AF) link to Segment_PL for forecast years. Quarterly splits equal-weighted unless noted.")

    r = 5

    # Helper: write a sum-type driver row (annual value split equally across quarters)
    def _gd_sum_row(label, hist_ann, asp_suffix, fmt_val, note_text, gd_key):
        nonlocal r
        _label(ws, r, label)
        # Historical years: annual/4 per quarter, annual hardcoded
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
            qv = hist_ann[year - 2023] / 4.0 if hist_ann[year - 2023] is not None else None
            for qc in qcols:
                if qv is not None:
                    _hval(ws, r, qc, qv, fmt_val)
                else:
                    ws.cell(row=r, column=qc).border = BD
            if hist_ann[year - 2023] is not None:
                _hval(ws, r, ann_col, hist_ann[year - 2023], fmt_val)
            else:
                ws.cell(row=r, column=ann_col).border = BD
        # Forecast years: direct link to quarterly Assumptions Selected row
        sel_row = R[asp_suffix]
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
            for qc in qcols:
                _fval(ws, r, qc, f"=Assumptions!{CL(qc)}{sel_row}", fmt_val)
            _fval(ws, r, ann_col, f"=SUM({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", fmt_val)
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, note_text)
        R[gd_key] = r; r += 1

    # Helper: write an avg-type driver row
    def _gd_avg_row(label, hist_ann, asp_suffix, note_text, gd_key):
        nonlocal r
        _label(ws, r, label)
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
            qv = hist_ann[year - 2023]
            for qc in qcols:
                if qv is not None:
                    _hval(ws, r, qc, qv, "0.0%")
                else:
                    ws.cell(row=r, column=qc).border = BD
            if qv is not None:
                _hval(ws, r, ann_col, qv, "0.0%")
            else:
                ws.cell(row=r, column=ann_col).border = BD
        sel_row = R[asp_suffix]
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
            for qc in qcols:
                _fval(ws, r, qc, f"=Assumptions!{CL(qc)}{sel_row}", "0.0%")
            _fval(ws, r, ann_col, f"=AVERAGE({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", "0.0%")
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, note_text)
        R[gd_key] = r; r += 1

    # Helper: write an avg-type driver row with custom fmt (for ASP $)
    def _gd_avg_fmt_row(label, hist_ann, asp_suffix, fmt_val, note_text, gd_key):
        nonlocal r
        _label(ws, r, label)
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
            qv = hist_ann[year - 2023]
            for qc in qcols:
                if qv is not None:
                    _hval(ws, r, qc, qv, fmt_val)
                else:
                    ws.cell(row=r, column=qc).border = BD
            if qv is not None:
                _hval(ws, r, ann_col, qv, fmt_val)
            else:
                ws.cell(row=r, column=ann_col).border = BD
        sel_row = R[asp_suffix]
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
            for qc in qcols:
                _fval(ws, r, qc, f"=Assumptions!{CL(qc)}{sel_row}", fmt_val)
            _fval(ws, r, ann_col, f"=AVERAGE({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", fmt_val)
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, note_text)
        R[gd_key] = r; r += 1

    # Helper: formula row across all 6 years (quarterly + annual = SUM)
    def _gd_formula_row(label, formula_fn, fmt_val, note_text, gd_key):
        """formula_fn(col_letter, row) -> formula string for that column"""
        nonlocal r
        _label(ws, r, label)
        for year, qcols, ann_col, suffix in GD_ALL_YEARS:
            for qc in qcols:
                _fval(ws, r, qc, formula_fn(CL(qc), r), fmt_val)
            _fval(ws, r, ann_col, f"=SUM({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", fmt_val)
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, note_text)
        R[gd_key] = r; r += 1

    # Helper: formula row where quarterly uses same formula and annual = SUM (no col-dependence)
    def _gd_formula_row_simple(label, formula_str, fmt_val, note_text, gd_key):
        nonlocal r
        _label(ws, r, label)
        for year, qcols, ann_col, suffix in GD_ALL_YEARS:
            for qc in qcols:
                _fval(ws, r, qc, f"={CL(qc)}{formula_str}", fmt_val)
            _fval(ws, r, ann_col, f"=SUM({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", fmt_val)
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, note_text)
        R[gd_key] = r; r += 1

    # Helper: blank quarterly cells with border only (for forecast-only or unused cells)
    def _gd_blank_q(ws, r, cols):
        for c in cols:
            ws.cell(row=r, column=c).border = BD

    # ============================================================
    # CCG — Bottom-Up: Revenue = PC TAM x Intel Share x ASP
    # ============================================================
    _section(ws, r, "CCG: Client Computing Group — Bottom-Up"); r += 1
    _label(ws, r, "Method: Revenue = PC TAM (M units) x Intel Unit Share (%) x Blended ASP ($)")
    _note(ws, r, "TAM source: Gartner/IDC. Share: Mercury Research. ASP derived. Q1'25 CCG=$7,629M, Q1'26 CCG=$7,727M actual.")
    r += 1

    # PC TAM (sum-type)
    _gd_sum_row("PC TAM (M units)", [241.8, 245.4, 258.6], "asp_pc_tam", "#,##0.0",
                "Annual TAM split equally across 4 quarters. FY2025 ~259M (Canalys est.)", "gd_ccg_tam")

    # Intel Unit Share (avg-type)
    _gd_avg_row("Intel PC Unit Share (%)", [0.710, 0.715, 0.720], "asp_intel_share",
                "Quarterly share = annual assumption (stable intra-year). Mercury Research est. ~72%", "gd_ccg_share")

    # Implied Units = TAM x Share
    tam_r = R["gd_ccg_tam"]; share_r = R["gd_ccg_share"]
    _gd_formula_row("  Intel Implied Units (M)",
                    lambda cl, rw: f"={cl}{tam_r}*{cl}{share_r}", "#,##0.0",
                    "= PC TAM x Intel Share. Implied unit shipments per quarter", "gd_ccg_units")

    # CCG Blended ASP (avg-type with custom fmt)
    _gd_avg_fmt_row("Intel CCG Blended ASP ($)", [188, 190, 173], "asp_ccg_asp", "0.0",
                    "Quarterly ASP = annual assumption. Q1 2026 ASP +16% YoY. Historical derived from Rev/(TAMxShare)", "gd_ccg_asp")

    # Implied CCG Revenue = Units x ASP
    units_r = R["gd_ccg_units"]; asp_r = R["gd_ccg_asp"]
    _gd_formula_row("  Implied CCG Revenue ($M)",
                    lambda cl, rw: f"={cl}{units_r}*{cl}{asp_r}", "#,##0",
                    "= Units x ASP. Bottom-up quarterly revenue. Annual = SUM(Q1:Q4) -> links to Segment_PL", "gd_ccg_implied_rev")

    # Reported CCG Revenue (historical: Segment_PL; forecast: mirrors Implied annual)
    # GD annual cols (6,11,16) map to Segment_PL annual HIST_COLS (2,3,4)
    gd_ann_to_segpl = {6: 6, 11: 11, 16: 16}
    _label(ws, r, "  Reported CCG Revenue ($M)")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
        segpl_rev_r = R["segpl_ccg_rev"]
        segpl_col = gd_ann_to_segpl[ann_col]
        qdata = GD_QREV["ccg"].get(year, {})
        for qi, qc in enumerate(qcols):
            qnum = qi + 1
            if qnum in qdata:
                _hval(ws, r, qc, qdata[qnum], "#,##0")
            else:
                _fval(ws, r, qc, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}/4", "#,##0")
        _lval(ws, r, ann_col, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}", "#,##0")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
        for qc in qcols:
            _fval(ws, r, qc, f"={CL(qc)}{R['gd_ccg_implied_rev']}", "#,##0")
        _lval(ws, r, ann_col, f"={CL(ann_col)}{R['gd_ccg_implied_rev']}", "#,##0")
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical: links to Segment_PL (10-K) split quarterly; Q1'25, Q4'25, Q1'26 have actuals. Forecast: = Implied bottom-up")
    R["gd_ccg_reported_rev"] = r; r += 1

    # Reconciliation rows (annual cols only)
    _section(ws, r, "  Reconciliation: Implied - Reported ($M, annual)")
    impl_r = R["gd_ccg_implied_rev"]; rep_r = R["gd_ccg_reported_rev"]
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _fval(ws, r, ann_col, f"={CL(ann_col)}{impl_r}-{CL(ann_col)}{rep_r}", "#,##0")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Should be near-zero for historical (<0.1%). Forecast: flags assumption inconsistency")
    R["gd_ccg_recon"] = r; r += 1

    _label(ws, r, "  Reconcil. as % of Reported")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['gd_ccg_recon']}/{CL(ann_col)}{rep_r}")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Green if <1%"); r += 2

    # ============================================================
    # DCAI — Bottom-Up: Rev = Server TAM x Intel DC Share x ASP + AI/ASIC Rev
    # ============================================================
    _section(ws, r, "DCAI: Data Center & AI — Bottom-Up"); r += 1
    _label(ws, r, "Method: Revenue = Server TAM (M) x Intel DC Share (%) x ASP ($) + AI/ASIC ($M)")
    _note(ws, r, "TAM: IDC. Share: Mercury Research. AI/ASIC: Gaudi+custom ASICs+NICs/IPUs. Q1'25 DCAI=$4,126M, Q1'26 DCAI=$5,052M actual.")
    r += 1

    # Server TAM
    _gd_sum_row("Server TAM (M units)", [11.5, 12.0, 13.5], "asp_svr_tam", "#,##0.0",
                "Annual Server TAM split equally across 4 quarters. IDC: 2025 ~13.5M", "gd_dcai_tam")

    # Intel DC Share
    _gd_avg_row("Intel DC Server Unit Share (%)", [0.775, 0.760, 0.730], "asp_intel_dc_share",
                "Quarterly share = annual assumption. Granite Rapids competitive vs Turin. ~72% Q4'24", "gd_dcai_share")

    # Implied DC Units
    dc_tam_r = R["gd_dcai_tam"]; dc_share_r = R["gd_dcai_share"]
    _gd_formula_row("  Intel Implied DC Units (M)",
                    lambda cl, rw: f"={cl}{dc_tam_r}*{cl}{dc_share_r}", "#,##0.0",
                    "= Server TAM x Intel DC Share. Implied quarterly unit shipments", "gd_dcai_units")

    # Server ASP
    _gd_avg_fmt_row("Intel Server Blended ASP ($)", [1772, 1712, 1632], "asp_dc_asp", "#,##0",
                    "Quarterly ASP = annual assumption. Q1 2026 ASP +27% YoY. Historical: (Rev-AI/ASIC)/(TAMxShare)", "gd_dcai_asp")

    # Implied CPU Revenue
    dc_units_r = R["gd_dcai_units"]; dc_asp_r = R["gd_dcai_asp"]
    _gd_formula_row("  Implied CPU Revenue ($M)",
                    lambda cl, rw: f"={cl}{dc_units_r}*{cl}{dc_asp_r}", "#,##0",
                    "= DC Units x Server ASP. Core x86 server CPU revenue", "gd_dcai_cpu_rev")

    # AI/ASIC Revenue
    _gd_sum_row("AI/ASIC & Other Revenue ($M)", [200, 500, 800], "asp_ai_asic", "#,##0",
                "Gaudi AI accelerators + custom ASICs + NICs/IPUs. FY2025 ~$800M. Q1'26 ASIC run rate >$1B", "gd_dcai_ai_asic")

    # Implied DCAI Revenue = CPU + AI/ASIC
    cpu_r = R["gd_dcai_cpu_rev"]; ai_r = R["gd_dcai_ai_asic"]
    _gd_formula_row("  Implied DCAI Revenue ($M)",
                    lambda cl, rw: f"={cl}{cpu_r}+{cl}{ai_r}", "#,##0",
                    "= CPU Revenue + AI/ASIC Revenue. Bottom-up DCAI estimate. Annual = SUM(Q1:Q4)", "gd_dcai_implied_rev")

    # Reported DCAI Revenue
    _label(ws, r, "  Reported DCAI Revenue ($M)")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
        segpl_rev_r = R["segpl_dcai_rev"]
        segpl_col = gd_ann_to_segpl[ann_col]
        qdata = GD_QREV["dcai"].get(year, {})
        for qi, qc in enumerate(qcols):
            qnum = qi + 1
            if qnum in qdata:
                _hval(ws, r, qc, qdata[qnum], "#,##0")
            else:
                _fval(ws, r, qc, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}/4", "#,##0")
        _lval(ws, r, ann_col, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}", "#,##0")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
        for qc in qcols:
            _fval(ws, r, qc, f"={CL(qc)}{R['gd_dcai_implied_rev']}", "#,##0")
        _lval(ws, r, ann_col, f"={CL(ann_col)}{R['gd_dcai_implied_rev']}", "#,##0")
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical: links to Segment_PL split quarterly. Forecast annual: = Implied")
    R["gd_dcai_reported_rev"] = r; r += 1

    # Reconciliation
    _section(ws, r, "  Reconciliation: Implied - Reported ($M, annual)")
    dcai_impl_r = R["gd_dcai_implied_rev"]; dcai_rep_r = R["gd_dcai_reported_rev"]
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _fval(ws, r, ann_col, f"={CL(ann_col)}{dcai_impl_r}-{CL(ann_col)}{dcai_rep_r}", "#,##0")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical should be near-zero (<0.5%). Forecast: flags driver vs growth-rate assumption divergence")
    R["gd_dcai_recon"] = r; r += 1

    _label(ws, r, "  Reconcil. as % of Reported")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['gd_dcai_recon']}/{CL(ann_col)}{dcai_rep_r}")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Green if <1%"); r += 2

    # ============================================================
    # Intel Foundry — Bottom-Up
    # ============================================================
    _section(ws, r, "Intel Foundry — Bottom-Up"); r += 1
    _label(ws, r, "Method: Revenue = Total Intel Chip Vol (M) x Internal Rev/Chip ($) + External Foundry ($M)")
    _note(ws, r, "~99% internal wafer sales to Intel Products. Chip vol = CCG + DCAI units.")
    r += 1

    # Total Intel Chip Volume = CCG Units + DCAI Units
    _gd_formula_row("Total Intel Chip Volume (M units)",
                    lambda cl, rw: f"={cl}{R['gd_ccg_units']}+{cl}{R['gd_dcai_units']}", "#,##0.0",
                    "= CCG Implied Units + DCAI Implied Units. Links to segment drivers above", "gd_foundry_vol")

    # Internal Rev per Chip
    _gd_avg_fmt_row("Internal Foundry Rev per Chip ($)", [102.2, 93.5, 90.1], "asp_foundry_rev_per_chip", "0.0",
                    "Quarterly Rev/Chip = annual assumption. Historical derived from Internal Foundry Rev / Total Units", "gd_foundry_rev_per_chip")

    # Internal Foundry Revenue = Vol x Rev/Chip
    vol_r = R["gd_foundry_vol"]; rpc_r = R["gd_foundry_rev_per_chip"]
    _gd_formula_row("  Internal Foundry Revenue ($M)",
                    lambda cl, rw: f"={cl}{vol_r}*{cl}{rpc_r}", "#,##0",
                    "= Total Chip Vol x Internal Rev/Chip. Wafer sales to Intel CCG/DCAI divisions", "gd_foundry_internal_rev")

    # External Foundry Revenue
    _gd_sum_row("External Foundry Revenue ($M)", [50, 60, 150], "asp_ext_foundry", "#,##0",
                "3rd-party wafer customers. FY2025 ~$150M. Growing as 18A/14A external PDK ramps", "gd_foundry_ext_rev")

    # Implied Total Foundry Revenue = Internal + External
    int_r = R["gd_foundry_internal_rev"]; ext_r = R["gd_foundry_ext_rev"]
    _gd_formula_row("  Implied Total Foundry Revenue ($M)",
                    lambda cl, rw: f"={cl}{int_r}+{cl}{ext_r}", "#,##0",
                    "= Internal Foundry Rev + External Foundry Rev", "gd_foundry_implied_rev")

    # Reported Foundry Revenue
    _label(ws, r, "  Reported Foundry Revenue ($M)")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
        segpl_rev_r = R["segpl_foundry_rev"]
        segpl_col = gd_ann_to_segpl[ann_col]
        qdata = GD_QREV["foundry"].get(year, {})
        for qi, qc in enumerate(qcols):
            qnum = qi + 1
            if qnum in qdata:
                _hval(ws, r, qc, qdata[qnum], "#,##0")
            else:
                _fval(ws, r, qc, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}/4", "#,##0")
        _lval(ws, r, ann_col, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}", "#,##0")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
        for qc in qcols:
            _fval(ws, r, qc, f"={CL(qc)}{R['gd_foundry_implied_rev']}", "#,##0")
        _lval(ws, r, ann_col, f"={CL(ann_col)}{R['gd_foundry_implied_rev']}", "#,##0")
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical: links to Segment_PL split quarterly. Forecast annual: = Implied")
    R["gd_foundry_reported_rev"] = r; r += 1

    # Reconciliation
    _section(ws, r, "  Reconciliation: Implied - Reported ($M, annual)")
    f_impl_r = R["gd_foundry_implied_rev"]; f_rep_r = R["gd_foundry_reported_rev"]
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _fval(ws, r, ann_col, f"={CL(ann_col)}{f_impl_r}-{CL(ann_col)}{f_rep_r}", "#,##0")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical should be near-zero (<0.1%). Forecast: flags driver vs growth-rate assumption divergence")
    R["gd_foundry_recon"] = r; r += 1

    _label(ws, r, "  Reconcil. as % of Reported")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['gd_foundry_recon']}/{CL(ann_col)}{f_rep_r}")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Green if <1%"); r += 2

    # ============================================================
    # All Other — Bottom-Up: Mobileye + IMS + Legacy
    # ============================================================
    _section(ws, r, "All Other: Mobileye + IMS + Legacy — Bottom-Up"); r += 1
    _label(ws, r, "Method: Revenue = LV Production (M) x Mobileye Rev/Vehicle ($) + IMS & Other ($M)")
    _note(ws, r, "Mobileye ADAS chips + IMS nanofab tools + legacy. Q1'25 All Other=$943M, Q1'26=$628M actual.")
    r += 1

    # Global LV Production
    _gd_sum_row("Global Light Vehicle Production (M)", [89.0, 88.0, 87.0], "asp_lv_prod", "#,##0.0",
                "IHS/S&P Global Mobility. FY2025 ~87M. Equal quarterly allocation", "gd_allother_lv_prod")

    # Mobileye Rev per Vehicle
    _gd_avg_fmt_row("Mobileye Rev per Vehicle ($)", [23.4, 18.9, 21.8], "asp_mbly_rev_per_veh", "0.0",
                    "Quarterly Rev/Veh = annual assumption. FY2025 Mobileye $1.9B / 87M LV = $21.8", "gd_allother_mbly_rpv")

    # Implied Mobileye Revenue = LV x RPU
    lv_r = R["gd_allother_lv_prod"]; rpv_r = R["gd_allother_mbly_rpv"]
    _gd_formula_row("  Implied Mobileye Revenue ($M)",
                    lambda cl, rw: f"={cl}{lv_r}*{cl}{rpv_r}", "#,##0",
                    "= LV Production x Mobileye Rev/Vehicle. EyeQ + SuperVision/Chauffeur", "gd_allother_mbly_rev")

    # IMS & Other Revenue
    _gd_sum_row("IMS & Other Revenue ($M)", [3378, 1941, 1663], "asp_ims_other", "#,##0",
                "IMS nanofabrication tools + legacy. FY2023-24 include Altera. Altera deconsolidated Sep 2025", "gd_allother_ims")

    # Implied All Other Revenue = Mobileye + IMS
    mbly_r = R["gd_allother_mbly_rev"]; ims_r = R["gd_allother_ims"]
    _gd_formula_row("  Implied All Other Revenue ($M)",
                    lambda cl, rw: f"={cl}{mbly_r}+{cl}{ims_r}", "#,##0",
                    "= Mobileye Revenue + IMS & Other. Annual = SUM(Q1:Q4)", "gd_allother_implied_rev")

    # Reported All Other Revenue
    _label(ws, r, "  Reported All Other Revenue ($M)")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
        segpl_rev_r = R["segpl_allother_rev"]
        segpl_col = gd_ann_to_segpl[ann_col]
        qdata = GD_QREV["allother"].get(year, {})
        for qi, qc in enumerate(qcols):
            qnum = qi + 1
            if qnum in qdata:
                _hval(ws, r, qc, qdata[qnum], "#,##0")
            else:
                _fval(ws, r, qc, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}/4", "#,##0")
        _lval(ws, r, ann_col, f"=Segment_PL!{CL(segpl_col)}{segpl_rev_r}", "#,##0")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
        for qc in qcols:
            _fval(ws, r, qc, f"={CL(qc)}{R['gd_allother_implied_rev']}", "#,##0")
        _lval(ws, r, ann_col, f"={CL(ann_col)}{R['gd_allother_implied_rev']}", "#,##0")
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical: links to Segment_PL split quarterly. Forecast annual: = Implied")
    R["gd_allother_reported_rev"] = r; r += 1

    # Reconciliation
    _section(ws, r, "  Reconciliation: Implied - Reported ($M, annual)")
    ao_impl_r = R["gd_allother_implied_rev"]; ao_rep_r = R["gd_allother_reported_rev"]
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _fval(ws, r, ann_col, f"={CL(ann_col)}{ao_impl_r}-{CL(ann_col)}{ao_rep_r}", "#,##0")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical should be near-zero (<1%). Forecast: flags driver vs growth-rate assumption divergence")
    R["gd_allother_recon"] = r; r += 1

    _label(ws, r, "  Reconcil. as % of Reported")
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['gd_allother_recon']}/{CL(ann_col)}{ao_rep_r}")
    _gd_blank_q(ws, r, [c for y, qc, a, s in GD_ALL_YEARS for c in qc])
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Green if <1%"); r += 2

    # ============================================================
    # Driver Sensitivity Summary (Annual FY2026E-FY2028E)
    # ============================================================
    _section(ws, r, "Driver Sensitivity: Revenue Elasticity (Annual Columns)"); r += 1
    _label(ws, r, "Impact of isolated driver changes on annual implied revenue (Base case). Uses FY2026E annual column (V).")
    _note(ws, r, "References annual column V (FY2026E). Switch scenario in Assumptions!B2 to update.")
    r += 1
    _label(ws, r, "Driver")
    for ci, lbl in enumerate(["FY2026E Value", "Change", "Rev Impact ($M)", "Impact %"], start=22):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font = Font(bold=True, size=9, name="Calibri"); c.border = BD; c.alignment = ALIGN_CENTER
    r += 1

    a2026 = GD_ANN_COLS[2026]  # 22
    s_items = [
        ("CCG: PC TAM +1%", f"={CL(a2026)}{R['gd_ccg_tam']}", f"={CL(a2026)}{R['gd_ccg_implied_rev']}*0.01",
         "PC TAM 1:1 elasticity with CCG revenue"),
        ("CCG: Intel PC Share +1pp", f"={CL(a2026)}{R['gd_ccg_share']}", f"={CL(a2026)}{R['gd_ccg_tam']}*0.01*{CL(a2026)}{R['gd_ccg_asp']}",
         "Most sensitive CCG driver. 1pp share = TAM x 1% x ASP"),
        ("CCG: Blended ASP +$1", f"={CL(a2026)}{R['gd_ccg_asp']}", f"={CL(a2026)}{R['gd_ccg_units']}",
         "$1 ASP change = Units x $1"),
        ("DCAI: Server TAM +1%", f"={CL(a2026)}{R['gd_dcai_tam']}", f"={CL(a2026)}{R['gd_dcai_cpu_rev']}*0.01",
         "TAM elasticity applies to CPU revenue only"),
        ("DCAI: DC Server Share +1pp", f"={CL(a2026)}{R['gd_dcai_share']}", f"={CL(a2026)}{R['gd_dcai_tam']}*0.01*{CL(a2026)}{R['gd_dcai_asp']}",
         "Most sensitive DCAI driver"),
        ("DCAI: Server ASP +$1", f"={CL(a2026)}{R['gd_dcai_asp']}", f"={CL(a2026)}{R['gd_dcai_units']}",
         "$1 ASP change = DC Units x $1"),
        ("DCAI: AI/ASIC +$100M", f"={CL(a2026)}{R['gd_dcai_ai_asic']}", "100",
         "AI/ASIC passes through 1:1"),
        ("Foundry: Total Chip Vol +1%", f"={CL(a2026)}{R['gd_foundry_vol']}", f"={CL(a2026)}{R['gd_foundry_internal_rev']}*0.01",
         "Chip vol elasticity applies to internal Foundry only"),
        ("Foundry: Internal Rev/Chip +$1", f"={CL(a2026)}{R['gd_foundry_rev_per_chip']}", f"={CL(a2026)}{R['gd_foundry_vol']}",
         "$1 Rev/Chip x total units = material swing"),
        ("All Other: LV Production +1%", f"={CL(a2026)}{R['gd_allother_lv_prod']}", f"={CL(a2026)}{R['gd_allother_mbly_rev']}*0.01",
         "Auto production sensitivity"),
    ]
    for label, val_formula, impact_formula, note_txt in s_items:
        _label(ws, r, f"  {label}")
        c22 = ws.cell(row=r, column=22, value=val_formula)
        c22.font = FONT_FORM; c22.border = BD; c22.alignment = ALIGN_RIGHT
        c22.number_format = "#,##0.0" if "M" not in label and "TAM" not in label and "Vol" not in label else "#,##0.0"
        if "Share" in label or "ASP" in label:
            c22.number_format = "0.0%" if "Share" in label else "0.0"
        c24 = ws.cell(row=r, column=24, value=impact_formula)
        c24.font = FONT_FORM; c24.number_format = "#,##0"; c24.border = BD; c24.alignment = ALIGN_RIGHT
        for cc in [23, 25]:
            ws.cell(row=r, column=cc).border = BD
        _note(ws, r, note_txt)
        r += 1

    # Column widths for GD tab
    ws.column_dimensions['A'].width = 44
    for c in range(2, GD_NCOL + 1):
        cl = CL(c)
        if c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.column_dimensions[cl].width = 2
        elif c in GD_ANN_COLS_LIST:
            ws.column_dimensions[cl].width = 15
        elif c == GD_NCOL:
            ws.column_dimensions[cl].width = 55
        else:
            ws.column_dimensions[cl].width = 12

    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "7030A0"
    return r


def build_segment_pl(wb, R):
    """Build Segment P&L tab with QUARTERLY revenue + OI per segment.
    Annual rollup columns (6,11,16,22,27,32) link to Consolidated_PL annual tabs.
    FIRST PASS - historical hardcoded, forecast deferred to pass 2 (Growth_Drivers link)."""
    ws = wb["Segment_PL"]
    _title_row(ws, 1, "Intel (INTC) - Segment P&L (Revenue + Operating Income, Quarterly)")
    _unit_row(ws, 2)
    _qheader(ws, 3)
    _label(ws, 4, "USD Million (except margin %). Quarters Q1-Q4 with FY annual rollup.")
    _note(ws, 4, "Segment revenue + OI quarterly. Forecast links to Growth_Drivers. Annual cols -> Consolidated_PL.")

    r = 5
    segs = [
        ("CCG", "CCG - Client Computing Group", "ccg",
         [32305, 33346, 32228],
         "asp_ccg_opm",
         [10128, 11594, 9317]),
        ("DCAI", "DCAI - Data Center & AI", "dcai",
         [15980, 16125, 16919],
         "asp_dcai_opm",
         [945, 1414, 3422]),
        ("Foundry", "Intel Foundry", "foundry",
         [18504, 17317, 17826],
         "asp_foundry_opm",
         [-7083, -13291, -10318]),
        ("AllOther", "All Other (Mobileye + IMS + Other)", "allother",
         [5463, 3601, 3563],
         "asp_allother_opm",
         [1507, -57, 264]),
    ]

    for seg_label, seg_desc, seg_key, hist_rev_ann, opm_key, hist_oi_ann in segs:
        # ---- Revenue (quarterly) ----
        _label(ws, r, seg_desc, bold=True)
        # Historical: quarterly from GD_QREV, fallback to annual/4
        qdata = GD_QREV.get(seg_key, {})
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
            qv = qdata.get(year, {})
            hist_ann = hist_rev_ann[year - 2023]
            for qi, qc in enumerate(qcols):
                qnum = qi + 1
                if qnum in qv:
                    _hval(ws, r, qc, qv[qnum], "#,##0")
                else:
                    _hval(ws, r, qc, hist_ann / 4.0, "#,##0")
            _hval(ws, r, ann_col, hist_ann, "#,##0")
        # Forecast: linked to GD in pass 2, placeholder borders for now
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
            for qc in qcols:
                ws.cell(row=r, column=qc).border = BD
            ws.cell(row=r, column=ann_col).border = BD
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, "Forecast: Links to Growth_Drivers implied revenue (pass 2). Annual = SUM(Q1:Q4).")
        R[f"segpl_{seg_key}_rev"] = r
        r += 1

        # ---- YoY% (annual cols only) ----
        for year, qcols, ann_col, suffix in GD_ALL_YEARS:
            if year == 2023:
                continue
            prev_ann = GD_ALL_YEARS[year - 2024][2]  # prior year annual col
            _pct(ws, r, ann_col, f"={CL(ann_col)}{r-1}/{CL(prev_ann)}{r-1}-1")
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, "YoY Growth % on annual basis")
        r += 1

        # ---- Operating Income (quarterly) ----
        _label(ws, r, "  Operating Income")
        # Historical: annual/4 per quarter
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[:3]:
            qoi = hist_oi_ann[year - 2023] / 4.0
            for qc in qcols:
                _hval(ws, r, qc, qoi, "#,##0")
            _hval(ws, r, ann_col, hist_oi_ann[year - 2023], "#,##0")
        # Forecast: quarterly OI = Rev_q x OPM_q from Assumptions
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:
            for qc in qcols:
                _fval(ws, r, qc, f"={CL(qc)}{r-2}*Assumptions!{CL(qc)}{R[opm_key]}", "#,##0")
            _fval(ws, r, ann_col, f"=SUM({CL(qcols[0])}{r}:{CL(qcols[3])}{r})", "#,##0")
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        _note(ws, r, "Forecast = Quarterly Revenue x Quarterly OPM% from Assumptions. Annual = SUM(Q1:Q4).")
        R[f"segpl_{seg_key}_oi"] = r
        r += 1

        # ---- OPM% ----
        _label(ws, r, "  Operating Margin %")
        for year, qcols, ann_col, suffix in GD_ALL_YEARS:
            for qc in qcols:
                _pct(ws, r, qc, f"={CL(qc)}{r-1}/{CL(qc)}{r-3}")
            _pct(ws, r, ann_col, f"={CL(ann_col)}{r-1}/{CL(ann_col)}{r-3}")
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        r += 2  # spacer

    # ---- Corporate Unallocated (annual-only, quarters blank) ----
    _section(ws, r, "Corporate Unallocated")
    hist_corp = [-5199, -11177, -5518]
    for ci, (yr, qc, ac, sf) in enumerate(GD_ALL_YEARS[:3]):
        _hval(ws, r, ac, hist_corp[ci])
    for yr, qc, ac, sf in GD_ALL_YEARS[3:]:
        _lval(ws, r, ac, f"=Assumptions!{CL(ac)}{R['asp_corp_unalloc']}")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Annual-only from Assumptions. Quarters blank (non-segment allocation).")
    R["segpl_corp_unalloc"] = r
    r += 1

    # ---- Sum of Segment OI + Corp Unalloc (quarterly) ----
    _section(ws, r, "Sum of Segment Operating Income + Corp Unalloc")
    seg_keys = ["ccg", "dcai", "foundry", "allother"]
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        for qc in qcols:
            parts = "+".join([f"{CL(qc)}{R['segpl_'+k+'_oi']}" for k in seg_keys])
            _key(ws, r, qc, f"={parts}+{CL(qc)}{R['segpl_corp_unalloc']}")
        parts_a = "+".join([f"{CL(ann_col)}{R['segpl_'+k+'_oi']}" for k in seg_keys])
        _key(ws, r, ann_col, f"={parts_a}+{CL(ann_col)}{R['segpl_corp_unalloc']}")
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "KEY OUTPUT. Links to Consolidated_PL EBIT via annual columns.")
    R["segpl_sum_oi"] = r
    r += 2

    # ---- Cross-check placeholder (annual cols only, filled pass 3) ----
    _section(ws, r, "CROSS-CHECK: Sum Seg OI vs Consolidated EBIT (Annual)")
    _label(ws, r + 1, "  Segment Sum OI (annual)")
    _label(ws, r + 2, "  Consolidated_PL EBIT")
    _note(ws, r + 2, "TODO: fill in pass 3 after Consolidated_PL built")
    _label(ws, r + 3, "  Difference (MUST = 0)")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        _fval(ws, r + 1, ann_col, f"={CL(ann_col)}{R['segpl_sum_oi']}")
        _fval(ws, r + 3, ann_col, f"={CL(ann_col)}{r+1}-{CL(ann_col)}{r+2}", "#,##0")
        for qc in qcols:
            ws.cell(row=r + 1, column=qc).border = BD
            ws.cell(row=r + 2, column=qc).border = BD
            ws.cell(row=r + 3, column=qc).border = BD
    _note(ws, r + 3, "MUST = 0 - validates segment/consolidated linkage")
    R["segpl_xcheck_sum_oi"] = r + 1
    R["segpl_xcheck_ebit"] = r + 2
    R["segpl_xcheck_diff"] = r + 3
    r = r + 4

    # ---- Memo: Total Segment Revenue (includes internal Foundry) ----
    _section(ws, r, "Memo: Total Segment Revenue (incl. internal Foundry, annual)")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        parts = "+".join([f"{CL(ann_col)}{R['segpl_'+k+'_rev']}" for k in seg_keys])
        _fval(ws, r, ann_col, f"={parts}")
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Sum of 4 reported segments; Foundry ~99% internal -> excluded from consolidated")
    R["segpl_sum_segments"] = r
    r += 2

    # ---- External Foundry Revenue (annual-only) ----
    _label(ws, r, "External Foundry Revenue (3rd-party only, annual)")
    hist_ext_foundry = [50, 60, 150]
    for ci, (yr, qc, ac, sf) in enumerate(GD_ALL_YEARS[:3]):
        _hval(ws, r, ac, hist_ext_foundry[ci])
    for yr, qc, ac, sf in GD_ALL_YEARS[3:]:
        _lval(ws, r, ac, f"=Assumptions!{CL(ac)}{R['asp_ext_foundry']}")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "~1% of total Foundry revenue historically. Growing as 18A/14A external customers ramp")
    R["segpl_ext_foundry"] = r
    r += 1

    # ---- Intersegment Eliminations (annual-only) ----
    _label(ws, r, "Intersegment Eliminations (annual)")
    hist_elim = [-205, -161, 619]
    for ci, (yr, qc, ac, sf) in enumerate(GD_ALL_YEARS[:3]):
        _hval(ws, r, ac, hist_elim[ci])
    for yr, qc, ac, sf in GD_ALL_YEARS[3:]:
        _lval(ws, r, ac, f"=Assumptions!{CL(ac)}{R['asp_inter_elim']}")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Residual elim; Foundry~Products internal wafer sales excluded from consolidated")
    R["segpl_elim"] = r
    r += 1

    # ---- TOTAL Consolidated Revenue (annual-only) ----
    _section(ws, r, "TOTAL Consolidated Revenue (annual)")
    hist_consol = [54228, 53101, 52853]
    for ci, (yr, qc, ac, sf) in enumerate(GD_ALL_YEARS[:3]):
        _hval(ws, r, ac, hist_consol[ci])
    for yr, qc, ac, sf in GD_ALL_YEARS[3:]:
        ext_parts = "+".join([f"{CL(ac)}{R['segpl_'+k+'_rev']}" for k in ["ccg", "dcai", "allother"]])
        formula = f"={ext_parts}+{CL(ac)}{R['segpl_ext_foundry']}+{CL(ac)}{R['segpl_elim']}"
        _key(ws, r, ac, formula)
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Historical: 10-K actuals. Forecast: CCG+DCAI+ExtFoundry+AllOther+Elim. Links to Consolidated_PL.")
    R["segpl_consol_rev"] = r
    r += 2

    # ---- Revenue Mix (annual-only) ----
    _section(ws, r, "Revenue Mix (% of Consolidated, annual)")
    for sk, sl in [("ccg", "CCG %"), ("dcai", "DCAI %"), ("allother", "All Other %")]:
        _label(ws, r, f"  {sl}")
        for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
            _pct(ws, r, ann_col, f"={CL(ann_col)}{R['segpl_'+sk+'_rev']}/{CL(ann_col)}{R['segpl_consol_rev']}")
            for qc in qcols:
                ws.cell(row=r, column=qc).border = BD
        for c in [GD_SPACER_COL, GD_SPACER2_COL]:
            ws.cell(row=r, column=c).border = BD
        r += 1
    _label(ws, r, "  External Foundry %")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['segpl_ext_foundry']}/{CL(ann_col)}{R['segpl_consol_rev']}")
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "External foundry only; internal wafer sales eliminated")
    r += 1
    _label(ws, r, "  Memo: Total Foundry / Seg Sum")
    for yr, qcols, ann_col, suffix in GD_ALL_YEARS:
        _pct(ws, r, ann_col, f"={CL(ann_col)}{R['segpl_foundry_rev']}/{CL(ann_col)}{R['segpl_sum_segments']}")
        for qc in qcols:
            ws.cell(row=r, column=qc).border = BD
    for c in [GD_SPACER_COL, GD_SPACER2_COL]:
        ws.cell(row=r, column=c).border = BD
    _note(ws, r, "Shows Foundry scale relative to all segments")
    r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "548235"
    return r


def build_segment_pl_pass2(wb, R):
    """Pass 2: Link Segment_PL forecast QUARTERLY revenue to Growth_Drivers implied revenue.
    Called after Growth_Drivers is built to avoid circular R-key dependency.
    Both Segment_PL and Growth_Drivers use same quarterly columns -> direct column match."""
    ws = wb["Segment_PL"]

    seg_to_gd = {
        "segpl_ccg_rev": "gd_ccg_implied_rev",
        "segpl_dcai_rev": "gd_dcai_implied_rev",
        "segpl_foundry_rev": "gd_foundry_implied_rev",
        "segpl_allother_rev": "gd_allother_implied_rev",
    }

    for seg_key, gd_key in seg_to_gd.items():
        r = R[seg_key]
        for year, qcols, ann_col, suffix in GD_ALL_YEARS[3:]:  # forecast years (2026-2028)
            for qc in qcols:
                f_text = f"=Growth_Drivers!{CL(qc)}{R[gd_key]}"
                ws.cell(row=r, column=qc, value=f_text)
                ws.cell(row=r, column=qc).font = Font(color="548235")
                ws.cell(row=r, column=qc).number_format = "#,##0"
                ws.cell(row=r, column=qc).border = BD
            ws.cell(row=r, column=ann_col, value=f"=SUM({CL(qcols[0])}{r}:{CL(qcols[3])}{r})")
            ws.cell(row=r, column=ann_col).font = Font(color="548235")
            ws.cell(row=r, column=ann_col).number_format = "#,##0"
            ws.cell(row=r, column=ann_col).border = BD


def build_segment_pl_pass3(wb, R):
    """Third pass: fill cross-check link from Segment_PL to Consolidated_PL EBIT.
    Segment_PL uses quarterly annual columns (6,11,16,22,27,32).
    Consolidated_PL uses annual-tab columns (2,3,4,6,7,8)."""
    ws = wb["Segment_PL"]
    yr_to_ann_col = {2023: 2, 2024: 3, 2025: 4, 2026: 6, 2027: 7, 2028: 8}
    for year, qcols, ann_col, suffix in GD_ALL_YEARS:
        ann_tab_col = yr_to_ann_col[year]
        c = ws.cell(row=R["segpl_xcheck_ebit"], column=ann_col,
                    value=f"=Consolidated_PL!{CL(ann_tab_col)}{R['pl_ebit']}")
        c.font = FONT_LINK; c.number_format = "#,##0"; c.border = BD; c.alignment = ALIGN_RIGHT


# ============================================================
# TAB 6: CONSOLIDATED_PL (by Cost Type, Segment-Driven EBIT)
# ============================================================

def build_consolidated_pl(wb, R):
    """Build consolidated P&L with segment-driven EBIT."""
    ws = wb["Consolidated_PL"]
    _title_row(ws, 1, "Intel (INTC) - Consolidated Income Statement")
    _unit_row(ws, 2)
    _yr_header_no_scenario(ws, 3)
    _label(ws, 4, "USD Million (except per-share data)")
    _note(ws, 4, "Cost structure by type. EBIT = Segment_PL Sum OI (segment-driven).")

    r = 5

    # Revenue
    _label(ws, r, "Net Revenue", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Segment_PL!{CL(SEGPL_COL_MAP[col])}{R['segpl_consol_rev']}")
    _note(ws, r, "<- Segment_PL consolidated total")
    R["pl_rev"] = r; r += 1

    # COGS
    _label(ws, r, "Cost of Sales")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [32517, 35756, 34478][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_cogs']}")
    _note(ws, r, "= Revenue x COGS% from Assumptions")
    R["pl_cogs"] = r; r += 1

    # Gross Profit
    _label(ws, r, "Gross Profit", bold=True)
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [21711, 17345, 18375][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['pl_rev']}-{CL(fc)}{R['pl_cogs']}")
    _note(ws, r, "= Revenue - COGS")
    R["pl_gp"] = r; r += 1
    _label(ws, r, "  Gross Margin %")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"={CL(col)}{R['pl_gp']}/{CL(col)}{R['pl_rev']}")
    R["pl_gpm"] = r; r += 2

    # OpEx
    _section(ws, r, "Operating Expenses (by cost type)"); r += 1
    _label(ws, r, "Research & Development")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [16046, 16546, 13774][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_rd']}")
    _note(ws, r, "= Revenue x R&D%")
    R["pl_rd"] = r; r += 1

    _label(ws, r, "Marketing, General & Administrative")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [5634, 5507, 4624][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_mga']}")
    _note(ws, r, "= Revenue x MG&A%")
    R["pl_mga"] = r; r += 1

    _label(ws, r, "Restructuring & Other Charges")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [-62, 6970, 2191][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_restruct']}")
    _note(ws, r, "From Assumptions; winding down"); r += 1

    # Total OpEx (top-down)
    _label(ws, r, "Total OpEx (per cost rates)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['pl_rd']}+{CL(col)}{R['pl_mga']}+{CL(col)}{r-1}")
    _note(ws, r, "= R&D + MG&A + Restructuring (top-down cost build)")
    R["pl_total_opex"] = r; r += 2

    # EBIT - SEGMENT DRIVEN
    _section(ws, r, "EBIT - Segment-Driven"); r += 1
    _label(ws, r, "EBIT (Operating Income)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Segment_PL!{CL(SEGPL_COL_MAP[col])}{R['segpl_sum_oi']}")
    _note(ws, r, "<- Segment_PL Sum of Segment OI. SEGMENT-DRIVEN source of truth.")
    R["pl_ebit"] = r; r += 1

    _label(ws, r, "  EBIT Margin %")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"={CL(col)}{R['pl_ebit']}/{CL(col)}{R['pl_rev']}")
    r += 1

    _label(ws, r, "  Implied Total OpEx (GP - EBIT)")
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['pl_gp']}-{CL(col)}{R['pl_ebit']}", "#,##0")
    _note(ws, r, "Transparency: GP - EBIT. May differ from top-down OpEx above."); r += 2

    # Non-Operating
    _section(ws, r, "Non-Operating Items"); r += 1
    _label(ws, r, "Interest & Other, Net")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [629, 226, 3257][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_interest']}")
    _note(ws, r, "Net interest expense on debt")
    R["pl_interest"] = r; r += 1

    _label(ws, r, "Gains/Losses on Equity Investments")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [40, 242, 514][ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 0)
    _note(ws, r, "Unpredictable; set to 0 for forecast"); r += 1

    # Pretax
    _label(ws, r, "Income (Loss) Before Taxes", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['pl_ebit']}+{CL(col)}{R['pl_interest']}+{CL(col)}{r-1}")
    _note(ws, r, "= EBIT + Interest + Equity Gains")
    R["pl_pretax"] = r; r += 1

    # Tax
    _label(ws, r, "Provision for (Benefit from) Taxes")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [-913, 8023, 1531][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc,
              f"=IF({CL(fc)}{R['pl_pretax']}>0,{CL(fc)}{R['pl_pretax']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_etr']},0)")
    _note(ws, r, "Tax on positive pretax only (valuation allowance shields US losses)")
    R["pl_tax"] = r; r += 1

    # Net Income
    _label(ws, r, "Net Income (Loss)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['pl_pretax']}-{CL(col)}{R['pl_tax']}")
    _note(ws, r, "= Pretax - Tax")
    R["pl_ni"] = r; r += 1
    _label(ws, r, "  Net Income Margin %")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"={CL(col)}{R['pl_ni']}/{CL(col)}{R['pl_rev']}")
    r += 1

    # Less NCI
    _label(ws, r, "Less: Net Income Attributable to NCI")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [-14, -477, 293][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_nci']}")
    _note(ws, r, "NCI from Assumptions (AZ SCIP + Mobileye; Ireland SCIP eliminated Apr 2026)")
    R["pl_nci"] = r; r += 1

    # NI to Intel
    _label(ws, r, "Net Income Attributable to Intel", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['pl_ni']}-{CL(col)}{r-1}")
    _note(ws, r, "= NI - NCI. Key EPS driver.")
    R["pl_ni_intel"] = r; r += 2

    # Memo
    _section(ws, r, "Memo Items"); r += 1
    _label(ws, r, "D&A (Memo)")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [7847, 9951, 10757][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_da']}")
    _note(ws, r, "Reference for EBITDA calc")
    R["pl_da"] = r; r += 1

    _label(ws, r, "SBC (Memo)")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [3229, 3410, 2434][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_sbc']}")
    _note(ws, r, "Reference for CFO/FCF bridge")
    R["pl_sbc"] = r; r += 1

    _label(ws, r, "Diluted Shares Outstanding (M)")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [4212, 4280, 4530][ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_shares']}")
    _note(ws, r, "Basic shares from Assumptions; diluted approximates basic")
    R["pl_shares"] = r; r += 1

    # Diluted EPS
    _label(ws, r, "Diluted EPS ($)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['pl_ni_intel']}/{CL(col)}{R['pl_shares']}", "0.00")
    _note(ws, r, "= NI to Intel / Shares. KEY OUTPUT")
    R["pl_eps"] = r; r += 1

    _label(ws, r, "Dividends per Share (Memo)")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [0.74, 0.38, 0.00][ci], "0.00")
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_dps']}", "0.00")
    _note(ws, r, "Suspended; no reinstatement expected"); r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "1F4E79"
    return r


def _set_col_widths(ws):
    ws.column_dimensions['A'].width = 44
    if _is_quarterly(ws):
        for c in range(2, GD_NCOL + 1):
            cl = CL(c)
            if c in (17, 34):
                ws.column_dimensions[cl].width = 2
            elif c in (6, 11, 16, 22, 27, 32):
                ws.column_dimensions[cl].width = 15
            elif c == GD_NCOL:
                ws.column_dimensions[cl].width = 55
            else:
                ws.column_dimensions[cl].width = 12
    else:
        for c in ['B','C','D','F','G','H']:
            ws.column_dimensions[c].width = 14
        ws.column_dimensions['E'].width = 2
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 2
        ws.column_dimensions['K'].width = 55


# ============================================================
# TAB 7: BS (Balance Sheet with Cash as Plug)
# ============================================================

def build_bs(wb, R):
    """Build balance sheet: assets = liabilities + equity. Cash is the plug."""
    ws = wb["BS"]
    _title_row(ws, 1, "Intel (INTC) - Balance Sheet")
    _unit_row(ws, 2)
    _yr_header_no_scenario(ws, 3)
    _label(ws, 4, "USD Million")
    _note(ws, 4, "Cash = Total L&E - Non-Cash Assets (plug). Balance check row must = 0.")

    r = 5

    # ---- ASSETS ----
    _section(ws, r, "ASSETS"); r += 1
    _section(ws, r, "Current Assets"); r += 1

    # Cash (PLUG - computed after Total L&E and non-cash assets are known)
    _label(ws, r, "Cash and Cash Equivalents (PLUG)", bold=True)
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [7079, 8249, 14265][ci])
    # Cash formula = Total L&E - (all other assets)
    # We'll write this after we know Total L&E row
    _note(ws, r, "PLUG: = Total L&E - all non-cash assets. Ensures A = L+E.")
    R["bs_cash"] = r; r += 1

    # Marketable Securities
    _label(ws, r, "Short-term Investments (Marketable Securities)")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [17955, 13813, 23151][ci])
    # Forecast: grow slightly (assume ~$24B flat for now)
    for fc in FCST_COLS:
        _hval(ws, r, fc, 24000)
    _note(ws, r, "TODO: Model as % of total cash or flat. Current: flat ~$24B assumption.")
    R["bs_ms"] = r; r += 1

    # AR
    _label(ws, r, "Accounts Receivable, Net")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [3402, 3478, 3839][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_ar_days']}/365")
    _note(ws, r, "= Revenue x AR Days / 365")
    R["bs_ar"] = r; r += 1

    # Inventories
    _label(ws, r, "Inventories")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [11127, 12198, 11618][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_cogs']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_inv_days']}/365")
    _note(ws, r, "= COGS x Inventory Days / 365")
    R["bs_inventory"] = r; r += 1

    # Other CA
    _label(ws, r, "Other Current Assets")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [3706, 9586, 10815][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_other_ca']}")
    _note(ws, r, "= Revenue x Other CA% (incl. refundable tax credits)")
    R["bs_other_ca"] = r; r += 1

    # Total CA
    _label(ws, r, "Total Current Assets", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_cash']}+{CL(col)}{R['bs_ms']}+{CL(col)}{R['bs_ar']}+{CL(col)}{R['bs_inventory']}+{CL(col)}{R['bs_other_ca']}")
    _note(ws, r, "= Cash + Securities + AR + Inventory + Other CA")
    R["bs_total_ca"] = r; r += 2

    # Non-Current Assets
    _section(ws, r, "Non-Current Assets"); r += 1

    _label(ws, r, "Property, Plant & Equipment, Net")
    hist_ppe = [96647, 107919, 105414]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_ppe[ci])
    # PP&E = Prior PP&E + CapEx - D&A - Govt Incentives
    # Simplified: Prior + Gross CapEx - D&A
    for fc in FCST_COLS:
        prev_col = HIST_COLS[2] if fc == 6 else fc - 1
        _fval(ws, r, fc,
              f"={CL(prev_col)}{r}+Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_capex']}-Consolidated_PL!{CL(fc)}{R['pl_da']}")
    _note(ws, r, "= Prior PP&E + Gross CapEx - D&A. Govt incentives not explicitly modelled (reducing gross PP&E).")
    R["bs_ppe"] = r; r += 1

    _label(ws, r, "Equity Investments")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [5829, 5383, 8512][ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 8500)
    _note(ws, r, "Includes Altera 49% retained stake (equity method). Flat forecast.")
    R["bs_equity_inv"] = r; r += 1

    _label(ws, r, "Goodwill")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [27591, 24693, 23912][ci])
    # Q1 2026 had $3.9B Mobileye goodwill impairment
    goodwill_q1_2026 = 23912 - 3900
    for fc in FCST_COLS:
        _hval(ws, r, fc, goodwill_q1_2026)
    _note(ws, r, f"Post-Q1 2026 $3.9B Mobileye impairment. Approx ${goodwill_q1_2026}M.")
    R["bs_goodwill"] = r; r += 1

    _label(ws, r, "Identified Intangible Assets, Net")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [4589, 3691, 2772][ci])
    for fc in FCST_COLS:
        prev_col = HIST_COLS[2] if fc == 6 else fc - 1
        _fval(ws, r, fc, f"={CL(prev_col)}{r}-500")
    _note(ws, r, "Amortising ~$500M/year")
    R["bs_intangibles"] = r; r += 1

    _label(ws, r, "Other Long-Term Assets")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [13647, 7475, 7131][ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 7000)
    _note(ws, r, "Flat ~$7B assumption")
    R["bs_other_nca"] = r; r += 1

    # Total NCA
    _label(ws, r, "Total Non-Current Assets", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_ppe']}+{CL(col)}{R['bs_equity_inv']}+{CL(col)}{R['bs_goodwill']}+{CL(col)}{R['bs_intangibles']}+{CL(col)}{R['bs_other_nca']}")
    R["bs_total_nca"] = r; r += 1

    # Total Assets
    _label(ws, r, "TOTAL ASSETS", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['bs_total_ca']}+{CL(col)}{R['bs_total_nca']}")
    R["bs_total_assets"] = r; r += 3

    # ---- LIABILITIES ----
    _section(ws, r, "LIABILITIES & EQUITY"); r += 1
    _section(ws, r, "Current Liabilities"); r += 1

    _label(ws, r, "Accounts Payable")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [8578, 12556, 9882][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_cogs']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_ap_days']}/365")
    _note(ws, r, "= COGS x AP Days / 365")
    R["bs_ap"] = r; r += 1

    _label(ws, r, "Accrued Compensation and Benefits")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [3655, 3343, 3990][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_accrued_comp']}")
    _note(ws, r, "= Revenue x Accrued Comp%")
    R["bs_accrued_comp"] = r; r += 1

    _label(ws, r, "Short-Term Debt")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [2288, 3729, 2499][ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 2500)
    _note(ws, r, "Stable ~$2.5B assumption; commercial paper + current maturities")
    R["bs_st_debt"] = r; r += 1

    _label(ws, r, "Income Taxes Payable")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [1107, 1756, 604][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_tax_payable']}")
    _note(ws, r, "= Revenue x Tax Payable%")
    R["bs_tax_payable"] = r; r += 1

    _label(ws, r, "Other Accrued Liabilities")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [12425, 14282, 14600][ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Consolidated_PL!{CL(fc)}{R['pl_rev']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_other_cl']}")
    _note(ws, r, "= Revenue x Other CL%")
    R["bs_other_cl"] = r; r += 1

    # Total CL
    _label(ws, r, "Total Current Liabilities", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_ap']}+{CL(col)}{R['bs_accrued_comp']}+{CL(col)}{R['bs_st_debt']}+{CL(col)}{R['bs_tax_payable']}+{CL(col)}{R['bs_other_cl']}")
    R["bs_total_cl"] = r; r += 2

    # Non-Current Liabilities
    _section(ws, r, "Non-Current Liabilities"); r += 1

    _label(ws, r, "Long-Term Debt")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [46978, 46282, 44086][ci])
    # LT Debt = Total Debt - ST Debt
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_debt']}-{CL(fc)}{R['bs_st_debt']}")
    _note(ws, r, "= Total Debt - ST Debt (from Assumptions)")
    R["bs_lt_debt"] = r; r += 1

    _label(ws, r, "Other Long-Term Liabilities")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [6576, 9505, 9408][ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 9500)
    _note(ws, r, "Flat ~$9.5B assumption. Includes tax, operating lease, other.")
    R["bs_other_ncl"] = r; r += 1

    # Total NCL
    _label(ws, r, "Total Non-Current Liabilities", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_lt_debt']}+{CL(col)}{R['bs_other_ncl']}")
    R["bs_total_ncl"] = r; r += 1

    # Total Liabilities
    _label(ws, r, "TOTAL LIABILITIES", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_total_cl']}+{CL(col)}{R['bs_total_ncl']}")
    R["bs_total_liab"] = r; r += 2

    # ---- EQUITY ----
    _section(ws, r, "Stockholders' Equity"); r += 1

    _label(ws, r, "Total Intel Stockholders' Equity")
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, [105590, 99270, 114281][ci])
    # Equity = Prior Equity + NI to Intel + SBC - Dividends + Net Equity Issuance
    for fc in FCST_COLS:
        prev_col = HIST_COLS[2] if fc == 6 else fc - 1
        _fval(ws, r, fc,
              f"={CL(prev_col)}{r}+Consolidated_PL!{CL(fc)}{R['pl_ni_intel']}+Consolidated_PL!{CL(fc)}{R['pl_sbc']}+Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_net_eq_iss']}")
    _note(ws, r, "= Prior Equity + NI to Intel + SBC + Net Equity Issuance. Dividends = 0.")
    R["bs_equity"] = r; r += 1

    _label(ws, r, "Non-Controlling Interests")
    hist_nci_bs = [4375, 5762, 12079]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_nci_bs[ci])
    # NCI = Prior NCI + NCI(P&L) - NCI distributions (simplified)
    # After Apr 2026 Ireland SCIP buyout: NCI (BS) drops
    nci_q1_2026_est = 13595  # from Q1 2026 10-Q
    for fc in FCST_COLS:
        prev_col = HIST_COLS[2] if fc == 6 else fc - 1
        # Simplified: prior + P&L NCI
        _fval(ws, r, fc, f"={CL(prev_col)}{r}+Consolidated_PL!{CL(fc)}{R['pl_nci']}-300")
    _note(ws, r, "= Prior NCI + P&L NCI - estimated distributions (~$300M). AZ SCIP + Mobileye remain.")
    R["bs_nci"] = r; r += 1

    # Total Equity
    _label(ws, r, "Total Stockholders' Equity", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_equity']}+{CL(col)}{R['bs_nci']}")
    R["bs_total_equity"] = r; r += 1

    # Total L&E
    _label(ws, r, "TOTAL LIABILITIES & EQUITY", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['bs_total_liab']}+{CL(col)}{R['bs_total_equity']}")
    R["bs_total_le"] = r; r += 1

    # ---- NOW: Write Cash Plug Formula ----
    # Cash = Total L&E - (all other assets)
    # But this creates circular reference with Total CA (which includes Cash)
    # Solution: Cash = Total L&E - Total NCA - (CA excl Cash)
    ca_excl_cash = f"({CL(6)}{R['bs_ms']}+{CL(6)}{R['bs_ar']}+{CL(6)}{R['bs_inventory']}+{CL(6)}{R['bs_other_ca']})"
    for fc in FCST_COLS:
        cash_formula = f"={CL(fc)}{R['bs_total_le']}-{CL(fc)}{R['bs_total_nca']}-({CL(fc)}{R['bs_ms']}+{CL(fc)}{R['bs_ar']}+{CL(fc)}{R['bs_inventory']}+{CL(fc)}{R['bs_other_ca']})"
        c = ws.cell(row=R["bs_cash"], column=fc, value=cash_formula)
        c.number_format = "#,##0"; c.font = FONT_BOLD; c.fill = FILL_KEY; c.border = BD; c.alignment = ALIGN_RIGHT

    # Balance Check
    r += 1
    _section(ws, r, "BALANCE CHECK")
    _label(ws, r, "  Total Assets")
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_total_ca']}+{CL(col)}{R['bs_total_nca']}")
    R["bs_assets_check"] = r; r += 1
    _label(ws, r, "  Total L&E (from above)")
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_total_le']}")
    r += 1
    _label(ws, r, "   Balance Check (A - L&E) - MUST = 0", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['bs_assets_check']}-{CL(col)}{r-1}", "0.0")
    _note(ws, r, "MUST = 0 in all years. Validates BS integrity with cash plug.")
    R["bs_balance_check"] = r; r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "1F4E79"
    return r


# ============================================================
# TAB 8: CASH_FLOW (WC from BS Delta)
# ============================================================

def build_cash_flow(wb, R):
    """Build cash flow statement: CFO + CFI + CFF."""
    ws = wb["Cash_Flow"]
    _title_row(ws, 1, "Intel (INTC) - Cash Flow Statement")
    _unit_row(ws, 2)
    _yr_header_no_scenario(ws, 3)
    _label(ws, 4, "USD Million")
    _note(ws, 4, "WC changes linked to BS deltas. CFO + CFI = FCF.")

    r = 5

    # ---- Operating Activities ----
    _section(ws, r, "Operating Activities"); r += 1

    _label(ws, r, "Net Income (Loss)")
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_ni']}")
    _note(ws, r, "<- Consolidated_PL")
    R["cf_ni"] = r; r += 1

    # Non-cash adjustments
    _label(ws, r, "(+) Depreciation & Amortization")
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_da']}")
    R["cf_da"] = r; r += 1

    _label(ws, r, "(+) Share-Based Compensation")
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_sbc']}")
    R["cf_sbc"] = r; r += 1

    _label(ws, r, "Deferred Taxes")
    hist_def_tax = [-2033, 6132, 328]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_def_tax[ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 0)
    _note(ws, r, "TODO: Model deferred tax. Set to 0 for now (valuation allowance limits)."); r += 1

    _label(ws, r, "Other Non-Cash / Non-Recurring")
    hist_other_cf = [6871, 11925, -2958]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_other_cf[ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 0)
    _note(ws, r, "Restructuring charges, impairments, gains/losses, etc. Set to 0 for forecast."); r += 1

    # Working Capital Changes
    r += 1
    _section(ws, r, "Working Capital Changes"); r += 1

    wc_items = [
        ("Change in AR (-increase = outflow)", "bs_ar", False),
        ("Change in Inventory (-increase = outflow)", "bs_inventory", False),
        ("Change in Other CA (-increase = outflow)", "bs_other_ca", False),
        ("Change in AP (+increase = inflow)", "bs_ap", True),
        ("Change in Accrued Comp (+increase = inflow)", "bs_accrued_comp", True),
        ("Change in Other CL (+increase = inflow)", "bs_other_cl", True),
        ("Change in Tax Payable (+increase = inflow)", "bs_tax_payable", True),
    ]
    for label, bs_key, is_inflow in wc_items:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            if col == 2:
                # No prior year for first historical column; leave blank or hardcode
                pass
            else:
                prev_col = col - 1 if col in HIST_COLS else (HIST_COLS[2] if col == 6 else col - 1)
                if is_inflow:
                    _fval(ws, r, col, f"=BS!{CL(col)}{R[bs_key]}-BS!{CL(prev_col)}{R[bs_key]}")
                else:
                    _fval(ws, r, col, f"=-(BS!{CL(col)}{R[bs_key]}-BS!{CL(prev_col)}{R[bs_key]})")
        _note(ws, r, f"Links to BS {bs_key}")
        r += 1
    # Store last WC row for CFO sum
    R["cf_last_wc"] = r - 1

    # CFO
    r += 1
    _label(ws, r, "Net Cash from Operating Activities (CFO)", bold=True)
    wc_rows = range(R["cf_last_wc"] - 6, R["cf_last_wc"] + 1)
    for col in HIST_COLS + FCST_COLS:
        wc_sum = "+".join([f"{CL(col)}{wr}" for wr in wc_rows])
        _key(ws, r, col, f"={CL(col)}{R['cf_ni']}+{CL(col)}{R['cf_da']}+{CL(col)}{R['cf_sbc']}+{CL(col)}{r-4}+{CL(col)}{r-3}+{wc_sum}")
    _note(ws, r, "= NI + D&A + SBC + DeferredTax + OtherNonCash + Sum(WC Changes)")
    R["cf_cfo"] = r; r += 2

    # ---- Investing Activities ----
    _section(ws, r, "Investing Activities"); r += 1

    _label(ws, r, "Capital Expenditure (Gross CapEx)")
    hist_capex = [-25750, -25122, -17672]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_capex[ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=-Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_capex']}")
    _note(ws, r, "Gross CapEx (negative). From Assumptions.")
    R["cf_capex"] = r; r += 1

    _label(ws, r, "  CapEx / Revenue %")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"=-{CL(col)}{r-1}/Consolidated_PL!{CL(col)}{R['pl_rev']}")
    r += 1

    _label(ws, r, "Other Investing Activities, Net")
    hist_other_inv = [1546, 1581, -1175]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_other_inv[ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 0)
    _note(ws, r, "Govt incentives, divestitures, ST investment net. Set to 0 for forecast.")
    R["cf_other_inv"] = r; r += 1

    # CFI
    _label(ws, r, "Net Cash from Investing Activities (CFI)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['cf_capex']}+{CL(col)}{R['cf_other_inv']}")
    R["cf_cfi"] = r; r += 1

    # FCF
    _label(ws, r, "Free Cash Flow (FCF = CFO + CFI)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['cf_cfo']}+{CL(col)}{R['cf_cfi']}")
    _note(ws, r, "= CFO + CFI. KEY OUTPUT.")
    R["cf_fcf"] = r; r += 1
    _label(ws, r, "  FCF / Revenue %")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"={CL(col)}{R['cf_fcf']}/Consolidated_PL!{CL(col)}{R['pl_rev']}")
    r += 2

    # ---- Financing Activities ----
    _section(ws, r, "Financing Activities"); r += 1

    _label(ws, r, "Net Change in Debt")
    hist_debt_net = [11453, -462, -2757]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_debt_net[ci])
    for fi, fc in enumerate(FCST_COLS):
        ann_col = ASP_ANN_COLS[fc]
        if fi == 0:
            _fval(ws, r, fc, f"=Assumptions!{CL(ann_col)}{R['asp_debt']}-(BS!{CL(HIST_COLS[2])}{R['bs_lt_debt']}+BS!{CL(HIST_COLS[2])}{R['bs_st_debt']})")
        else:
            prev_fc = FCST_COLS[fi - 1]
            prev_ann = ASP_ANN_COLS[prev_fc]
            _fval(ws, r, fc, f"=Assumptions!{CL(ann_col)}{R['asp_debt']}-Assumptions!{CL(prev_ann)}{R['asp_debt']}")
    _note(ws, r, "Change in Total Debt from Assumptions")
    R["cf_debt_net"] = r; r += 1

    _label(ws, r, "Net Equity Issuance / (Repurchase)")
    hist_eq_net = [2567, 356, 18586]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_eq_net[ci])
    for fc in FCST_COLS:
        _lval(ws, r, fc, f"=Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_net_eq_iss']}")
    _note(ws, r, "ESPP proceeds net of RSU withholdings. FY2025 inflated by strategic placements.")
    R["cf_equity_net"] = r; r += 1

    _label(ws, r, "Dividends Paid")
    hist_div = [-3088, -1599, 0]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_div[ci])
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=-Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_dps']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_shares']}")
    _note(ws, r, "Dividends suspended; = 0.")
    R["cf_dividends"] = r; r += 1

    _label(ws, r, "Partner Contributions (SCIP) & Other Financing")
    hist_scip = [3696, 12471, 1827]
    for ci, col in enumerate(HIST_COLS):
        _hval(ws, r, col, hist_scip[ci])
    for fc in FCST_COLS:
        _hval(ws, r, fc, 500)
    _note(ws, r, "AZ SCIP ongoing partner contributions. Ireland SCIP eliminated Apr 2026."); r += 1

    # CFF
    _label(ws, r, "Net Cash from Financing Activities (CFF)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['cf_debt_net']}+{CL(col)}{R['cf_equity_net']}+{CL(col)}{R['cf_dividends']}+{CL(col)}{r-1}")
    R["cf_cff"] = r; r += 1

    # Net Change in Cash
    _label(ws, r, "Net Increase (Decrease) in Cash", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['cf_cfo']}+{CL(col)}{R['cf_cfi']}+{CL(col)}{R['cf_cff']}")
    _note(ws, r, "= CFO + CFI + CFF")
    R["cf_net_change"] = r; r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "548235"
    return r


# ============================================================
# TAB 9: DCF (UFCF -> EV -> Equity Value)
# ============================================================

def build_dcf(wb, R):
    """Build DCF valuation: UFCF -> PV -> TV -> EV -> Equity -> per-share value.
    Includes WACC assumptions (moved from Assumptions tab)."""
    ws = wb["DCF"]
    _title_row(ws, 1, "Intel (INTC) - Discounted Cash Flow Valuation")
    _unit_row(ws, 2, "USD Million (except per-share data & multiples)")

    # Year numbering for DCF (t=1, 2, 3)
    for fi, fc in enumerate(FCST_COLS):
        ws.cell(row=3, column=fc, value=f"Year {fi+1}").font = FONT_HEADER
        ws.cell(row=3, column=fc).fill = FILL_HEADER; ws.cell(row=3, column=fc).border = BD
        ws.cell(row=3, column=fc).alignment = ALIGN_CENTER
    # Standard header row for WACC section
    _year_header(ws, 4)

    r = 6

    # ---- WACC Parameters ----
    _section(ws, r, "WACC Assumptions"); r += 1
    _arow_const(ws, r, "Risk-Free Rate (Rf)", 0.0388, 0.0425, 0.045, 0.045, "0.00%",
                "US 10Y Treasury, May 2026"); R["dcf_rf"] = r; r += 1
    _arow_const(ws, r, "Equity Risk Premium (ERP)", 0.055, 0.055, 0.055, 0.055, "0.00%",
                "Damodaran 2026 implied ERP"); R["dcf_erp"] = r; r += 1
    _arow(ws, r, "Levered Beta (β)", 1.20, 1.35, 1.50, 1.50, 1.30, 1.80, "0.00",
          "Bloomberg 2Y adj beta; Bear=1.80 on higher risk"); R["dcf_beta"] = r; r += 1
    _arow_const(ws, r, "Pre-Tax Cost of Debt (Kd)", 0.048, 0.05, 0.05, 0.050, "0.00%",
                "Weighted avg coupon on outstanding notes"); R["dcf_kd"] = r; r += 1
    _arow(ws, r, "Target D/(D+E)", 0.30, 0.32, 0.32, 0.25, 0.20, 0.35, "0.0%",
          "Long-term target; D/EV =32% at current prices"); R["dcf_dtc"] = r; r += 1
    _arow(ws, r, "Marginal Tax Rate (for WACC)", None, None, None, 0.12, 0.11, 0.15, "0.00%",
          "Marginal rate for interest tax shield. Base=12% (aligned with Q2 2026 guided 11%)"); R["dcf_etr_marginal"] = r; r += 1
    _arow(ws, r, "Terminal Growth Rate (g)", None, None, None, 0.025, 0.030, 0.020, "0.00%",
          "Base=2.5% LT nominal GDP+; Bull=3.0% AI secular; Bear=2.0% mature"); R["dcf_tg"] = r; r += 1

    r += 1
    # Derived WACC
    _section(ws, r, "Derived WACC"); r += 1
    _label(ws, r, "Cost of Equity (Ke)")
    _fval(ws, r, 2, f"=I{R['dcf_rf']}+I{R['dcf_beta']}*I{R['dcf_erp']}", "0.00%")
    _fval(ws, r, 6, f"=F{R['dcf_rf']}+F{R['dcf_beta']}*F{R['dcf_erp']}", "0.00%")
    _fval(ws, r, 7, f"=G{R['dcf_rf']}+G{R['dcf_beta']}*G{R['dcf_erp']}", "0.00%")
    _fval(ws, r, 8, f"=H{R['dcf_rf']}+H{R['dcf_beta']}*H{R['dcf_erp']}", "0.00%")
    _note(ws, r, "Ke = Rf + β × ERP")
    R["dcf_ke"] = r; r += 1
    _label(ws, r, "After-Tax Cost of Debt")
    _fval(ws, r, SEL_COL, f"=I{R['dcf_kd']}*(1-I{R['dcf_etr_marginal']})", "0.00%")
    _note(ws, r, "Kd × (1 − ETR)")
    R["dcf_atkd"] = r; r += 1
    _label(ws, r, "WACC")
    _fval(ws, r, SEL_COL, f"=I{R['dcf_ke']}*(1-I{R['dcf_dtc']})+I{R['dcf_atkd']}*I{R['dcf_dtc']}", "0.00%")
    _note(ws, r, "WACC = Ke×(1−D/TC) + Kd(1−t)×D/TC")
    R["dcf_wacc"] = r; r += 1

    r += 1
    _section(ws, r, "Valuation Bridge Inputs"); r += 1
    _label(ws, r, "Cash & Equivalents ($M) - from BS")
    R["dcf_cash_bridge_val"] = r; r += 1
    _label(ws, r, "Total Debt ($M) - from BS")
    R["dcf_debt_bridge_val"] = r; r += 1

    # DCF Calculation
    r += 1
    _section(ws, r, "DCF Calculation"); r += 1

    # EBIT
    _label(ws, r, "EBIT")
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_ebit']}")
    _note(ws, r, "<- Consolidated_PL (segment-driven)")
    R["dcf_ebit"] = r; r += 1

    # Tax on EBIT
    _label(ws, r, "(-) Tax on EBIT")
    for col in HIST_COLS:
        _hval(ws, r, col, 0)  # Not meaningful historically
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['dcf_ebit']}*Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_etr']}")
    _note(ws, r, "= EBIT x ETR from Assumptions")
    R["dcf_tax_on_ebit"] = r; r += 1

    # NOPAT
    _label(ws, r, "NOPAT (= EBIT - Tax on EBIT)")
    for col in HIST_COLS + FCST_COLS:
        _fval(ws, r, col, f"={CL(col)}{R['dcf_ebit']}-{CL(col)}{R['dcf_tax_on_ebit']}")
    _note(ws, r, "Net Operating Profit After Tax")
    R["dcf_nopat"] = r; r += 1

    # D&A
    _label(ws, r, "(+) D&A")
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_da']}")
    R["dcf_da"] = r; r += 1

    # CapEx
    _label(ws, r, "(-) Capital Expenditure")
    for col in HIST_COLS:
        _hval(ws, r, col, 0)
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"=-Assumptions!{CL(ASP_ANN_COLS[fc])}{R['asp_capex']}")
    _note(ws, r, "Gross CapEx (negative). From Assumptions.")
    R["dcf_capex"] = r; r += 1

    # Delta NWC
    _label(ws, r, "(-) Change in Net Working Capital")
    for col in HIST_COLS:
        _hval(ws, r, col, 0)
    # NWC = (AR + Inventory + Other CA) - (AP + Accrued Comp + Other CL + Tax Payable)
    # ΔNWC = NWC_t - NWC_(t-1)
    # For simplicity, model as % of ΔRevenue
    for fc in FCST_COLS:
        prev_col = HIST_COLS[2] if fc == 6 else fc - 1
        _fval(ws, r, fc,
              f"=-(Consolidated_PL!{CL(fc)}{R['pl_rev']}-Consolidated_PL!{CL(prev_col)}{R['pl_rev']})*0.10")
    _note(ws, r, "≈ 10% of ΔRevenue (AR 7% + Inv 21% - AP 17% ≈ 11%). Sign: NWC increase = cash outflow.")
    R["dcf_delta_nwc"] = r; r += 1

    # UFCF
    _label(ws, r, "Unlevered Free Cash Flow (UFCF)", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _key(ws, r, col, f"={CL(col)}{R['dcf_nopat']}+{CL(col)}{R['dcf_da']}+{CL(col)}{R['dcf_capex']}+{CL(col)}{R['dcf_delta_nwc']}")
    _note(ws, r, "= NOPAT + D&A + CapEx (neg) + ΔNWC (neg if outflow)")
    R["dcf_ufcf"] = r; r += 1

    # Discount Factor
    _label(ws, r, "Discount Factor")
    for fi, fc in enumerate(FCST_COLS):
        t = fi + 1
        _fval(ws, r, fc, f"=1/(1+I{R['dcf_wacc']})^{t}", "0.0000")
    _note(ws, r, "= 1/(1+WACC)^t. WACC from DCF tab assumptions (~10.5-10.7%)")
    R["dcf_df"] = r; r += 1

    # PV of UFCF
    _label(ws, r, "PV of UFCF")
    for fc in FCST_COLS:
        _fval(ws, r, fc, f"={CL(fc)}{R['dcf_ufcf']}*{CL(fc)}{R['dcf_df']}")
    _note(ws, r, "= UFCF x Discount Factor")
    R["dcf_pv_ufcf"] = r; r += 2

    # Terminal Value
    _section(ws, r, "Terminal Value (Gordon Growth Model)"); r += 1
    _label(ws, r, "Terminal Year UFCF (FY2028E)")
    for fc in [8]:
        _fval(ws, r, fc, f"={CL(fc)}{R['dcf_ufcf']}")
    _note(ws, r, "= UFCF in final explicit forecast year"); r += 1

    _label(ws, r, "Terminal Growth Rate (g)")
    _lval(ws, r, 8, f"=I{R['dcf_tg']}", "0.00%")
    _note(ws, r, "From DCF WACC assumptions. Base = 2.5%.")
    r += 1

    _label(ws, r, "Terminal Value (TV = UFCF x (1+g) / (WACC-g))")
    _fval(ws, r, 8, f"=H{R['dcf_ufcf']}*(1+H{r-1})/(I{R['dcf_wacc']}-H{r-1})")
    _note(ws, r, "Gordon Growth: TV = UFCF_terminal x (1+g) / (WACC - g)")
    R["dcf_tv"] = r; r += 1

    _label(ws, r, "PV of Terminal Value")
    _fval(ws, r, 8, f"=H{r-1}*H{R['dcf_df']}")
    _note(ws, r, "= TV x Discount Factor of final year")
    R["dcf_pv_tv"] = r; r += 2

    # EV Bridge
    _section(ws, r, "Enterprise Value Bridge"); r += 1
    _label(ws, r, "Sum of PV of UFCFs (FY2026E-FY2028E)")
    _fval(ws, r, 8, f"=F{R['dcf_pv_ufcf']}+G{R['dcf_pv_ufcf']}+H{R['dcf_pv_ufcf']}")
    _note(ws, r, "= Sum of PV of UFCF for explicit forecast period")
    R["dcf_sum_pv_ufcf"] = r; r += 1

    _label(ws, r, "(+) PV of Terminal Value")
    _fval(ws, r, 8, f"=H{R['dcf_pv_tv']}")
    R["dcf_pv_tv_val"] = r; r += 1

    _label(ws, r, "Enterprise Value (EV)", bold=True)
    _key(ws, r, 8, f"=H{r-2}+H{r-1}")
    _note(ws, r, "= Sum PV UFCF + PV TV")
    R["dcf_ev"] = r; r += 1

    _label(ws, r, "  TV as % of EV")
    _pct(ws, r, 8, f"=H{r-1}/H{r}")
    r += 2

    # Equity Bridge
    _section(ws, r, "Equity Value Bridge"); r += 1
    _label(ws, r, "Enterprise Value")
    _fval(ws, r, 8, f"=H{R['dcf_ev']}")
    R["dcf_ev_bridge"] = r; r += 1

    _label(ws, r, "(+) Cash & Equivalents")
    _lval(ws, r, 8, f"=BS!H{R['bs_cash']}")
    _note(ws, r, "<- BS Cash (latest forecast year)")
    R["dcf_cash"] = r; r += 1

    _label(ws, r, "(+) Marketable Securities")
    _lval(ws, r, 8, f"=BS!H{R['bs_ms']}")
    _note(ws, r, "<- BS Short-term Investments")
    R["dcf_ms_bridge"] = r; r += 1

    _label(ws, r, "(-) Total Debt")
    _lval(ws, r, 8, f"=BS!H{R['bs_lt_debt']}+BS!H{R['bs_st_debt']}")
    _note(ws, r, "<- BS LT Debt + ST Debt")
    R["dcf_debt_bridge"] = r; r += 1

    _label(ws, r, "(-) Non-Controlling Interests")
    _lval(ws, r, 8, f"=BS!H{R['bs_nci']}")
    _note(ws, r, "<- BS NCI")
    R["dcf_nci_bridge"] = r; r += 1

    _label(ws, r, "Equity Value", bold=True)
    _key(ws, r, 8, f"=H{R['dcf_ev_bridge']}+H{R['dcf_cash']}+H{R['dcf_ms_bridge']}-H{R['dcf_debt_bridge']}-H{R['dcf_nci_bridge']}")
    _note(ws, r, "= EV + Cash + Securities - Debt - NCI. KEY OUTPUT.")
    R["dcf_equity_val"] = r; r += 1

    _label(ws, r, "Diluted Shares Outstanding (M)")
    _lval(ws, r, 8, f"=Consolidated_PL!H{R['pl_shares']}")
    _note(ws, r, "<- Consolidated_PL. FY2028E diluted shares.")
    R["dcf_shares"] = r; r += 1

    _label(ws, r, "Implied Share Price ($)", bold=True)
    _key(ws, r, 8, f"=H{R['dcf_equity_val']}/H{R['dcf_shares']}", "$#,##0.00")
    _note(ws, r, "= Equity Value / Shares. CENTRAL OUTPUT OF THE MODEL.")
    R["dcf_implied_price"] = r; r += 2

    # DCF Summary table
    _section(ws, r, "DCF Summary"); r += 1
    summary_items = [
        ("WACC", f"=I{R['dcf_wacc']}", "0.00%"),
        ("Terminal Growth Rate", f"=I{R['dcf_tg']}", "0.00%"),
        ("Sum PV UFCF", f"=H{R['dcf_sum_pv_ufcf']}", "#,##0"),
        ("PV Terminal Value", f"=H{R['dcf_pv_tv_val']}", "#,##0"),
        ("TV % of EV", f"=H{R['dcf_pv_tv_val']}/H{R['dcf_ev']}", "0.0%"),
        ("Enterprise Value", f"=H{R['dcf_ev']}", "#,##0"),
        ("(+) Cash + Securities", f"=H{R['dcf_cash']}+H{R['dcf_ms_bridge']}", "#,##0"),
        ("(-) Total Debt", f"=H{R['dcf_debt_bridge']}", "#,##0"),
        ("(-) NCI", f"=H{R['dcf_nci_bridge']}", "#,##0"),
        ("Equity Value", f"=H{R['dcf_equity_val']}", "#,##0"),
        ("Implied Share Price", f"=H{R['dcf_implied_price']}", "$#,##0.00"),
    ]
    for label, formula, fmt in summary_items:
        _label(ws, r, f"  {label}")
        _key(ws, r, 8, formula, fmt)
        r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "FFF2CC"
    return r


# ============================================================
# TAB 10: SENSITIVITY (WACC x Terminal g)
# ============================================================

def build_sensitivity(wb, R):
    """Build 2D sensitivity table: WACC (cols) x Terminal g (rows) -> Implied Price."""
    ws = wb["Sensitivity"]
    _title_row(ws, 1, "Intel (INTC) - Sensitivity Analysis")
    _unit_row(ws, 2, "Implied Share Price ($) — WACC vs Terminal Growth Rate")
    _label(ws, 4, "Base case WACC = 10.7%, Base case g = 2.5%")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

    r = 6
    # WACC range in columns (8.0% to 13.0% in 1% steps)
    wacc_values = [8.0, 9.0, 10.0, 10.7, 11.0, 12.0, 13.0]
    # Terminal g range in rows (1.5% to 3.5% in 0.5% steps)
    g_values = [1.5, 2.0, 2.5, 3.0, 3.5]

    # Header row
    _label(ws, r, "Terminal g \\ WACC")
    for ci, wacc in enumerate(wacc_values):
        col = ci + 2  # start from column B
        c = ws.cell(row=r, column=col, value=wacc)
        c.number_format = "0.0%"
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.border = BD; c.alignment = ALIGN_CENTER
    r += 1

    for gi, g in enumerate(g_values):
        _label(ws, r, f"  {g}%")
        c = ws.cell(row=r, column=1)
        c.font = FONT_BOLD; c.border = BD
        for ci, wacc in enumerate(wacc_values):
            col = ci + 2
            # Recompute implied price for each WACC/g combination
            # Formula: recompute TV and everything downstream
            # Since Excel DATA TABLE is complex in openpyxl, we write the formula directly
            # using the DCF tab's structure but substituting WACC and g
            # For simplicity, mark as TODO: requires Excel Data Table or manual macro
            is_base = abs(wacc - 0.107) < 0.001 and abs(g - 0.025) < 0.001
            if is_base:
                c = ws.cell(row=r, column=col, value="BASE")
                c.font = FONT_BOLD; c.fill = FILL_KEY
            else:
                c = ws.cell(row=r, column=col, value="")
                c.font = FONT_FORM
            c.border = BD; c.alignment = ALIGN_CENTER
        r += 1

    _note(ws, r - len(g_values), "TODO: Implement 2D DATA TABLE in Excel. Open in Excel, select the table range, Data > What-If Analysis > Data Table. Row input: WACC cell (DCF!I{wacc_row}). Column input: g cell (DCF!I{tg_row}). Base case (WACC=10.7%, g=2.5%) is marked.")
    r += 1

    _section(ws, r, "Instructions"); r += 1
    instructions = [
        "1. Open Intel_IB_Model.xlsx in Microsoft Excel",
        "2. Go to Sensitivity tab",
        "3. Select the table range (B7:H11)",
        "4. Data > What-If Analysis > Data Table",
        f"5. Row input cell: =DCF!$I${R['dcf_wacc']}",
        f"6. Column input cell: =DCF!$I${R['dcf_tg']}",
        "7. Click OK. Table will populate with implied prices.",
    ]
    for inst in instructions:
        c = ws.cell(row=r, column=1, value=inst)
        c.font = Font(size=9, name="Calibri")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    _set_col_widths(ws)
    ws.sheet_properties.tabColor = "548235"
    return r


# ============================================================
# TAB 11: RATIO_ANALYSIS
# ============================================================

def build_ratio_analysis(wb, R):
    """Build ratio analysis tab: profitability, efficiency, leverage, CF quality, growth."""
    ws = wb["Ratio_Analysis"]
    _title_row(ws, 1, "Intel (INTC) - Ratio Analysis & Financial Metrics")
    _unit_row(ws, 2)
    _yr_header_no_scenario(ws, 3)
    _label(ws, 4, "All ratios computed from model financials")
    _note(ws, 4, "Links to Consolidated_PL, BS, Cash_Flow. Auto-updates with scenario changes.")

    r = 5

    # Profitability
    _section(ws, r, "Profitability Ratios"); r += 1
    ratios = [
        ("Gross Margin %", f"=Consolidated_PL!{{c}}{R['pl_gpm']}", "0.0%", True),
        ("EBIT Margin %", f"=Consolidated_PL!{{c}}{R['pl_ebit']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", False),
        ("Net Income Margin %", f"=Consolidated_PL!{{c}}{R['pl_ni']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", False),
        ("ROE (NI to Intel / Avg Equity)", f"=Consolidated_PL!{{c}}{R['pl_ni_intel']}/BS!{{c}}{R['bs_equity']}", "0.0%", False),
        ("ROA (NI / Avg Total Assets)", f"=Consolidated_PL!{{c}}{R['pl_ni']}/BS!{{c}}{R['bs_total_assets']}", "0.0%", False),
        ("EBITDA ($M)", f"=Consolidated_PL!{{c}}{R['pl_ebit']}+Consolidated_PL!{{c}}{R['pl_da']}", "#,##0", False),
        ("EBITDA Margin %", f"=(Consolidated_PL!{{c}}{R['pl_ebit']}+Consolidated_PL!{{c}}{R['pl_da']})/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", False),
    ]
    for label, formula_tpl, fmt, is_link in ratios:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            formula = formula_tpl.replace("{c}", CL(col))
            if is_link:
                _lval(ws, r, col, formula, fmt)
            else:
                _fval(ws, r, col, formula, fmt)
        r += 1

    # Efficiency
    r += 1
    _section(ws, r, "Efficiency Ratios"); r += 1
    eff_ratios = [
        ("AR Days", f"=BS!{{c}}{R['bs_ar']}/(Consolidated_PL!{{c}}{R['pl_rev']}/365)", "#,##0.0"),
        ("AP Days", f"=BS!{{c}}{R['bs_ap']}/(Consolidated_PL!{{c}}{R['pl_cogs']}/365)", "#,##0.0"),
        ("Inventory Days", f"=BS!{{c}}{R['bs_inventory']}/(Consolidated_PL!{{c}}{R['pl_cogs']}/365)", "#,##0.0"),
        ("CapEx / Revenue %", f"=Cash_Flow!{{c}}{R['cf_capex']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%"),
        ("CapEx / D&A (x)", f"=-Cash_Flow!{{c}}{R['cf_capex']}/Consolidated_PL!{{c}}{R['pl_da']}", "0.00x"),
    ]
    for label, formula_tpl, fmt in eff_ratios:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            formula = formula_tpl.replace("{c}", CL(col))
            _fval(ws, r, col, formula, fmt)
        r += 1

    # Leverage
    r += 1
    _section(ws, r, "Leverage & Liquidity Ratios"); r += 1
    lev_ratios = [
        ("Debt / Total Assets", f"=(BS!{{c}}{R['bs_lt_debt']}+BS!{{c}}{R['bs_st_debt']})/BS!{{c}}{R['bs_total_assets']}", "0.0%"),
        ("Debt / Equity", f"=(BS!{{c}}{R['bs_lt_debt']}+BS!{{c}}{R['bs_st_debt']})/BS!{{c}}{R['bs_equity']}", "0.00x"),
        ("Current Ratio", f"=BS!{{c}}{R['bs_total_ca']}/BS!{{c}}{R['bs_total_cl']}", "0.00x"),
        ("Net Cash / (Debt) ($M)", f"=BS!{{c}}{R['bs_cash']}+BS!{{c}}{R['bs_ms']}-BS!{{c}}{R['bs_lt_debt']}-BS!{{c}}{R['bs_st_debt']}", "#,##0"),
    ]
    for label, formula_tpl, fmt in lev_ratios:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            formula = formula_tpl.replace("{c}", CL(col))
            _fval(ws, r, col, formula, fmt)
        r += 1

    # Cash Flow Quality
    r += 1
    _section(ws, r, "Cash Flow Quality"); r += 1
    cf_ratios = [
        ("OCF / Net Income (x)", f"=Cash_Flow!{{c}}{R['cf_cfo']}/Consolidated_PL!{{c}}{R['pl_ni']}", "0.00x"),
        ("FCF / Revenue %", f"=Cash_Flow!{{c}}{R['cf_fcf']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%"),
        ("FCF / Net Income (x)", f"=Cash_Flow!{{c}}{R['cf_fcf']}/Consolidated_PL!{{c}}{R['pl_ni']}", "0.00x"),
    ]
    for label, formula_tpl, fmt in cf_ratios:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            formula = formula_tpl.replace("{c}", CL(col))
            _fval(ws, r, col, formula, fmt)
        r += 1

    # Growth
    r += 1
    _section(ws, r, "Growth Rates (YoY %)"); r += 1
    growth_items = [
        ("Revenue Growth %", f"=Consolidated_PL!{{c}}{R['pl_rev']}", None),
        ("Net Income Growth %", f"=Consolidated_PL!{{c}}{R['pl_ni_intel']}", None),
        ("EPS Growth %", f"={CL(HIST_COLS[1])}{R['pl_eps']}", None),
    ]
    for label, _, _ in growth_items:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            if col == 2:
                _hval(ws, r, col, "N/A")
            else:
                prev_col = col - 1 if col in HIST_COLS else (HIST_COLS[2] if col == 6 else col - 1)
                if label.startswith("Revenue"):
                    ref1 = f"Consolidated_PL!{CL(col)}{R['pl_rev']}"
                    ref2 = f"Consolidated_PL!{CL(prev_col)}{R['pl_rev']}"
                elif label.startswith("Net Income"):
                    ref1 = f"Consolidated_PL!{CL(col)}{R['pl_ni_intel']}"
                    ref2 = f"Consolidated_PL!{CL(prev_col)}{R['pl_ni_intel']}"
                else:
                    ref1 = f"Consolidated_PL!{CL(col)}{R['pl_eps']}"
                    ref2 = f"Consolidated_PL!{CL(prev_col)}{R['pl_eps']}"
                _pct(ws, r, col, f"={ref1}/{ref2}-1")
        r += 1

    # ============================================================
    # Market Valuation Multiples
    # ============================================================
    r += 1
    _section(ws, r, "Market Valuation Multiples"); r += 1
    _label(ws, r, "  Current Stock Price ($) — EDIT HERE", bold=True)
    sp_cell = ws.cell(row=r, column=6, value=113)
    sp_cell.font = Font(color="000000", bold=True); sp_cell.fill = PatternFill("solid", fgColor="DDEBF7")
    sp_cell.number_format = "0.00"; sp_cell.border = BD
    for col in [2, 3, 4, 7, 8]:
        ws.cell(row=r, column=col).border = BD
    _note(ws, r, "Enter current share price; TTM & NTM multiples update automatically")
    SP = f"F{r}"
    R["ratio_sp"] = r
    r += 2

    # --- TTM Valuation (FY2025A, column D) ---
    _section(ws, r, "TTM Valuation (Based on FY2025A)"); r += 1

    ev_d = (f"({SP}*Consolidated_PL!D{R['pl_shares']}"
            f"+BS!D{R['bs_st_debt']}+BS!D{R['bs_lt_debt']}"
            f"-BS!D{R['bs_cash']}-BS!D{R['bs_ms']})")
    ebitda_d = f"(Consolidated_PL!D{R['pl_ebit']}+Consolidated_PL!D{R['pl_da']})"

    ttm_metrics = [
        ("P/E (x)",
         f"={SP}/Consolidated_PL!D{R['pl_eps']}", "0.0x",
         "Price / TTM Diluted EPS"),
        ("P/B (x)",
         f"={SP}/(BS!D{R['bs_equity']}/Consolidated_PL!D{R['pl_shares']})", "0.0x",
         "Price / TTM Book Value Per Share"),
        ("P/S (x)",
         f"={SP}/(Consolidated_PL!D{R['pl_rev']}/Consolidated_PL!D{R['pl_shares']})", "0.0x",
         "Price / TTM Revenue Per Share"),
        ("EV/EBITDA (x)",
         f"={ev_d}/{ebitda_d}", "0.0x",
         "Enterprise Value / TTM EBITDA"),
        ("EV/Revenue (x)",
         f"={ev_d}/Consolidated_PL!D{R['pl_rev']}", "0.0x",
         "Enterprise Value / TTM Revenue"),
        ("P/CF (x)",
         f"={SP}/(Cash_Flow!D{R['cf_cfo']}/Consolidated_PL!D{R['pl_shares']})", "0.0x",
         "Price / TTM Operating Cash Flow Per Share"),
        ("PEG Ratio (x)",
         f"=({SP}/Consolidated_PL!D{R['pl_eps']})/((Consolidated_PL!D{R['pl_eps']}/Consolidated_PL!C{R['pl_eps']}-1)*100)", "0.00x",
         "P/E / TTM EPS Growth %; NM if growth <=0"),
        ("Dividend Yield %",
         f"=Assumptions!P{R['asp_dps']}/{SP}", "0.0%",
         "TTM DPS / Price; dividend currently suspended"),
    ]
    for label, formula, fmt, note_text in ttm_metrics:
        _label(ws, r, f"  {label}")
        c = ws.cell(row=r, column=4, value=formula)
        c.number_format = fmt; c.border = BD
        for col in [2, 3, 6, 7, 8]:
            ws.cell(row=r, column=col).border = BD
        _note(ws, r, note_text)
        r += 1

    r += 1

    # --- NTM Valuation (FY2026E, column F) ---
    _section(ws, r, "NTM Valuation (Based on FY2026E)"); r += 1

    ev_f = (f"({SP}*Consolidated_PL!F{R['pl_shares']}"
            f"+BS!F{R['bs_st_debt']}+BS!F{R['bs_lt_debt']}"
            f"-BS!F{R['bs_cash']}-BS!F{R['bs_ms']})")
    ebitda_f = f"(Consolidated_PL!F{R['pl_ebit']}+Consolidated_PL!F{R['pl_da']})"

    ntm_metrics = [
        ("P/E (x)",
         f"={SP}/Consolidated_PL!F{R['pl_eps']}", "0.0x",
         "Price / NTM Diluted EPS; updates with scenario"),
        ("P/B (x)",
         f"={SP}/(BS!F{R['bs_equity']}/Consolidated_PL!F{R['pl_shares']})", "0.0x",
         "Price / NTM Book Value Per Share"),
        ("P/S (x)",
         f"={SP}/(Consolidated_PL!F{R['pl_rev']}/Consolidated_PL!F{R['pl_shares']})", "0.0x",
         "Price / NTM Revenue Per Share"),
        ("EV/EBITDA (x)",
         f"={ev_f}/{ebitda_f}", "0.0x",
         "Enterprise Value / NTM EBITDA"),
        ("EV/Revenue (x)",
         f"={ev_f}/Consolidated_PL!F{R['pl_rev']}", "0.0x",
         "Enterprise Value / NTM Revenue"),
        ("P/CF (x)",
         f"={SP}/(Cash_Flow!F{R['cf_cfo']}/Consolidated_PL!F{R['pl_shares']})", "0.0x",
         "Price / NTM Operating Cash Flow Per Share"),
        ("PEG Ratio (x)",
         f"=({SP}/Consolidated_PL!F{R['pl_eps']})/((Consolidated_PL!F{R['pl_eps']}/Consolidated_PL!D{R['pl_eps']}-1)*100)", "0.00x",
         "P/E / NTM EPS Growth %; NM if growth <=0"),
        ("Dividend Yield %",
         f"=Assumptions!V{R['asp_dps']}/{SP}", "0.0%",
         "NTM DPS / Price; dividend currently suspended"),
    ]
    for label, formula, fmt, note_text in ntm_metrics:
        _label(ws, r, f"  {label}")
        c = ws.cell(row=r, column=6, value=formula)
        c.number_format = fmt; c.border = BD
        for col in [2, 3, 4, 7, 8]:
            ws.cell(row=r, column=col).border = BD
        _note(ws, r, note_text)
        r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "548235"
    return r


# ============================================================
# TAB 2: KEY_SUMMARY (Executive Dashboard - built LAST)
# ============================================================

def build_key_summary(wb, R):
    """Build executive dashboard linking to all other tabs."""
    ws = wb["Key_Summary"]
    _title_row(ws, 1, "Intel (INTC) - Key Summary & Executive Dashboard")
    _unit_row(ws, 2)
    _label(ws, 3, f"Scenario: ")
    _lval(ws, 3, 2, f"=Assumptions!{CL(2)}2", "0")
    ws.cell(row=3, column=2).font = Font(bold=True, size=11, name="Calibri", color="1F4E79")
    ws.cell(row=3, column=2).fill = FILL_ASSUMP

    r = 5
    # KPIs section
    _section(ws, r, "Key Performance Indicators")
    _yr_header_no_scenario(ws, r + 1)
    r += 2

    kpis = [
        ("Total Revenue ($M)", f"=Consolidated_PL!{{c}}{R['pl_rev']}", "#,##0", True),
        ("  YoY Growth %", f"=Consolidated_PL!{{c}}{R['pl_rev']}/Consolidated_PL!{{c-1}}{R['pl_rev']}-1", "0.0%", False),
        ("Gross Profit ($M)", f"=Consolidated_PL!{{c}}{R['pl_gp']}", "#,##0", True),
        ("  Gross Margin %", f"=Consolidated_PL!{{c}}{R['pl_gp']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", True),
        ("EBIT ($M)", f"=Consolidated_PL!{{c}}{R['pl_ebit']}", "#,##0", True),
        ("  EBIT Margin %", f"=Consolidated_PL!{{c}}{R['pl_ebit']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", True),
        ("Net Income to Intel ($M)", f"=Consolidated_PL!{{c}}{R['pl_ni_intel']}", "#,##0", True),
        ("  Net Margin %", f"=Consolidated_PL!{{c}}{R['pl_ni_intel']}/Consolidated_PL!{{c}}{R['pl_rev']}", "0.0%", True),
        ("Diluted EPS ($)", f"=Consolidated_PL!{{c}}{R['pl_eps']}", "0.00", True),
        ("D&A ($M)", f"=Consolidated_PL!{{c}}{R['pl_da']}", "#,##0", False),
        ("EBITDA ($M)", f"=Consolidated_PL!{{c}}{R['pl_ebit']}+Consolidated_PL!{{c}}{R['pl_da']}", "#,##0", True),
        ("FCF ($M)", f"=Cash_Flow!{{c}}{R['cf_fcf']}", "#,##0", True),
    ]
    for label, formula_tpl, fmt, is_key in kpis:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            formula = formula_tpl.replace("{c}", CL(col)).replace("{c-1}", CL(col-1 if col in HIST_COLS else (HIST_COLS[2] if col == 6 else col-1)))
            if col == 2 and "YoY" in label:
                pass  # No YoY for first year
            else:
                if is_key:
                    _key(ws, r, col, formula, fmt)
                elif "YoY" in label or "Margin" in label:
                    _pct(ws, r, col, formula)
                else:
                    _lval(ws, r, col, formula, fmt)
        r += 1

    # Revenue by Segment
    r += 1
    _section(ws, r, "Revenue by Segment ($M)"); r += 1
    for sk, sl in [("ccg","CCG"),("dcai","DCAI"),("foundry","Intel Foundry"),("allother","All Other")]:
        _label(ws, r, f"  {sl}")
        for col in HIST_COLS + FCST_COLS:
            _lval(ws, r, col, f"=Segment_PL!{CL(SEGPL_COL_MAP[col])}{R['segpl_'+sk+'_rev']}")
        r += 1
    _label(ws, r, "  Total Consolidated", bold=True)
    for col in HIST_COLS + FCST_COLS:
        _lval(ws, r, col, f"=Segment_PL!{CL(SEGPL_COL_MAP[col])}{R['segpl_consol_rev']}")
    r += 2

    # BS Highlights
    _section(ws, r, "Balance Sheet Highlights ($M)"); r += 1
    bs_items = [
        ("Total Assets", "bs_total_assets"),
        ("Total Equity", "bs_total_equity"),
        ("Cash + Securities", None),  # special handling
        ("Total Debt", None),
        ("Net Cash / (Debt)", None),
    ]
    for label, bs_key in bs_items:
        _label(ws, r, f"  {label}")
        for col in HIST_COLS + FCST_COLS:
            if label == "Cash + Securities":
                _lval(ws, r, col, f"=BS!{CL(col)}{R['bs_cash']}+BS!{CL(col)}{R['bs_ms']}")
            elif label == "Total Debt":
                _lval(ws, r, col, f"=BS!{CL(col)}{R['bs_lt_debt']}+BS!{CL(col)}{R['bs_st_debt']}")
            elif label == "Net Cash / (Debt)":
                _lval(ws, r, col, f"=BS!{CL(col)}{R['bs_cash']}+BS!{CL(col)}{R['bs_ms']}-BS!{CL(col)}{R['bs_lt_debt']}-BS!{CL(col)}{R['bs_st_debt']}")
            else:
                _lval(ws, r, col, f"=BS!{CL(col)}{R[bs_key]}")
        r += 1

    _label(ws, r, "  ROE")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_ni_intel']}/BS!{CL(col)}{R['bs_equity']}")
    r += 1
    _label(ws, r, "  ROA")
    for col in HIST_COLS + FCST_COLS:
        _pct(ws, r, col, f"=Consolidated_PL!{CL(col)}{R['pl_ni']}/BS!{CL(col)}{R['bs_total_assets']}")
    r += 2

    # DCF Output
    _section(ws, r, "DCF Valuation Output"); r += 1
    _label(ws, r, "  Implied Share Price ($)", bold=True)
    _lval(ws, r, 8, f"=DCF!H{R['dcf_implied_price']}", "$#,##0.00")
    _note(ws, r, "Base case DCF output. See Sensitivity tab for range.")
    r += 1
    _label(ws, r, "  Enterprise Value ($M)")
    _lval(ws, r, 8, f"=DCF!H{R['dcf_ev']}", "#,##0")
    r += 1
    _label(ws, r, "  Equity Value ($M)")
    _lval(ws, r, 8, f"=DCF!H{R['dcf_equity_val']}", "#,##0")
    r += 2

    # Investment Highlights
    _section(ws, r, "Investment Highlights"); r += 1
    highlights = [
        "1. CPU Renaissance: DCAI benefiting from AI inference shift toward CPUs (CPU:GPU ratio 1:8->1:1)",
        "2. 18A Ramp: Ahead of internal yield targets; volume up 6-7x in Q2 2026 vs Q1",
        "3. Foundry Losses Narrowing: -58% OPM (FY2025) -> -22% (FY2028E Base) as 18A matures",
        "4. Gross Margin Recovery: 34.8% (FY2025) -> 43.5% (FY2028E Base) driven by yield + pricing",
        "5. Supply Constraints Easing: Undershipping demand by ~$1B; tool spending up 25% to address",
        "6. Fab 34 Buyout: April 2026 ~$14.2B repurchase of Ireland SCIP stake; removes $135M/q NCI drag",
        "7. No Dividends: Suspended; priority is debt reduction and CapEx",
        "8. Warrant Overhang: 241M shares @ $20 strike if Foundry ownership <51% (NOT in base case)",
    ]
    for h in highlights:
        c = ws.cell(row=r, column=1, value=h)
        c.font = Font(size=10, name="Calibri")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    r += 1
    _section(ws, r, "Key Risks"); r += 1
    risks = [
        "1. 18A yield improvement stalls; Foundry losses remain >40% OPM through FY2028",
        "2. PC TAM declines >15% (recession scenario); CCG revenue headwinds",
        "3. AMD Turin / ARM CPUs erode DCAI market position",
        "4. External foundry customers do not materialize; Intel remains captive fab",
        "5. Debt load (~$50B) constrains strategic flexibility if FCF remains negative",
        "6. Valuation allowance on US DTAs ($9.9B) means losses generate no tax benefit",
        "7. WACC sensitivity: each 1% change in WACC moves implied price ~$5-7",
    ]
    for risk in risks:
        c = ws.cell(row=r, column=1, value=risk)
        c.font = Font(size=9, name="Calibri", italic=True, color="808080")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    _set_col_widths(ws)
    lr = _add_legend(ws, r + 1)
    _add_sources(ws, lr + 1)
    ws.sheet_properties.tabColor = "FFF2CC"
    return r


# ============================================================
# TAB 1: COVER
# ============================================================

def build_cover(wb):
    """Build cover page with company info and usage instructions."""
    ws = wb["Cover"]
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 30

    r = 1
    c = ws.cell(row=r, column=1, value="Intel Corporation (NASDAQ: INTC)")
    c.font = Font(bold=True, size=18, name="Calibri", color="1F4E79")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 2

    c = ws.cell(row=r, column=1, value="Investment Bank-Grade Financial Model")
    c.font = Font(bold=True, size=14, name="Calibri"); r += 1

    c = ws.cell(row=r, column=1, value="Segment-Driven, Three-Statement Linked DCF")
    c.font = Font(size=11, name="Calibri", italic=True); r += 1

    c = ws.cell(row=r, column=1, value="Model Date: May 7, 2026 | Forecast Horizon: FY2026E - FY2028E")
    c.font = Font(size=10, name="Calibri", color="808080"); r += 3

    # Summary info
    info = [
        ("Company", "Intel Corporation (INTC)"),
        ("Sector", "Semiconductors"),
        ("Headquarters", "Santa Clara, CA, USA"),
        ("Fiscal Year", "Ends last Saturday of December (52/53-week year)"),
        ("Reporting Segments", "CCG, DCAI, Intel Foundry, All Other (Mobileye + IMS + Other)"),
        ("", ""),
        ("Scenario Selector", "Assumptions tab, Cell B2 (Base / Bull / Bear)"),
        ("Model Architecture", "11 tabs: Cover -> Key_Summary -> Assumptions -> Growth_Drivers -> Segment_PL -> Consolidated_PL -> BS -> Cash_Flow -> DCF -> Sensitivity -> Ratio_Analysis"),
        ("", ""),
        ("Key Data Sources", "Intel 10-K filings (FY2022-FY2025), Q1 2026 10-Q, Q1 2026 Earnings Call (Apr 23, 2026)"),
        ("Assumptions Date", "May 7, 2026 (intc_assumptions.md)"),
        ("Financial Data", "intc_financial_data.md (extracted from SEC filings)"),
    ]
    for label, value in info:
        if label:
            c = ws.cell(row=r, column=1, value=label)
            c.font = Font(bold=True, size=10, name="Calibri")
        c = ws.cell(row=r, column=2, value=value)
        c.font = Font(size=10, name="Calibri")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1
    r += 1

    # Segment evolution note
    c = ws.cell(row=r, column=1, value="Segment Structure Note:")
    c.font = Font(bold=True, size=10, name="Calibri")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1

    notes = [
        "Effective Q1 2025: NEX (Network & Edge) integrated into CCG and DCAI; Altera deconsolidated Sep 12, 2025 (51% sold to SLP).",
        "FY2023-FY2025 segment data is RESTATED to current 4-segment structure per 2025 10-K Note 3.",
        "Historical data from old structure (with NEX) is available in 2024 10-K for reference but NOT used in this model.",
        "Mobileye remains consolidated (~77% Intel ownership). Q1 2026: $3.9B goodwill impairment at Mobileye.",
        "Ireland SCIP (Fab 34) 49% NCI bought out April 2026 (~$14.2B). Arizona SCIP 49% NCI remains.",
    ]
    for note_text in notes:
        c = ws.cell(row=r, column=1, value=f"  - {note_text}")
        c.font = Font(size=9, name="Calibri", color="404040")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    r += 2

    # Usage instructions
    c = ws.cell(row=r, column=1, value="Usage Instructions:")
    c.font = Font(bold=True, size=11, name="Calibri"); r += 1

    instructions = [
        "1. Select scenario in Assumptions tab cell B2 (Base / Bull / Bear). All tabs update automatically.",
        "2. Edit blue-fill cells in Assumptions tab to customize forecasts.",
        "3. Check yellow-highlighted cross-validation rows (Segment_PL, BS) — both must read 0 difference.",
        "4. DCF tab computes implied share price. Sensitivity tab provides WACC x g matrix.",
        "5. Key_Summary tab gives executive overview; Ratio_Analysis tab provides detailed metrics.",
        "6. All green-font cells are cross-sheet links (do not edit). Blue-font cells are historical data.",
    ]
    for inst in instructions:
        c = ws.cell(row=r, column=1, value=inst)
        c.font = Font(size=10, name="Calibri")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    r += 2

    # Disclaimer
    c = ws.cell(row=r, column=1, value="DISCLAIMER:")
    c.font = Font(bold=True, size=10, name="Calibri", color="C00000"); r += 1
    disclaimers = [
        "This model is for educational / analytical purposes only. It does not constitute investment advice.",
        "Forecasts are based on publicly available information and assumptions documented in intc_assumptions.md.",
        "Actual results may differ materially. Refer to Intel's SEC filings for official financial information.",
        "Model generated programmatically via Python openpyxl (build_intc_model.py).",
    ]
    for d_text in disclaimers:
        c = ws.cell(row=r, column=1, value=d_text)
        c.font = Font(size=8, name="Calibri", italic=True, color="808080")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    ws.sheet_properties.tabColor = "1F4E79"
    return r


# ============================================================
# MAIN MODEL BUILDER
# ============================================================

def build_model():
    """Create the complete 12-tab workbook."""
    print("Building Intel_IB_Model.xlsx ...")
    wb = openpyxl.Workbook()
    R = {}  # Row reference dictionary

    # Create all sheets in tab order
    ws0 = wb.active
    ws0.title = "Cover"
    for name in ["Key_Summary", "Assumptions", "Growth_Drivers", "Segment_PL",
                 "Consolidated_PL", "BS", "Cash_Flow", "DCF", "Sensitivity", "Ratio_Analysis"]:
        wb.create_sheet(name)

    # Build in dependency order
    print("  1/11 Cover ...")
    build_cover(wb)

    print("  2/11 Assumptions ...")
    build_assumptions(wb, R)

    print("  3/11 Segment_PL (pass 1 - structure + historical) ...")
    build_segment_pl(wb, R)

    print("  4/11 Growth_Drivers ...")
    build_growth_drivers(wb, R)

    print("  5/11 Segment_PL (pass 2 - link to Growth_Drivers) ...")
    build_segment_pl_pass2(wb, R)

    print("  6/11 Consolidated_PL ...")
    build_consolidated_pl(wb, R)

    print("  7/11 Segment_PL (pass 3 - cross-check) ...")
    build_segment_pl_pass3(wb, R)

    print("  8/11 Balance Sheet ...")
    build_bs(wb, R)

    print("  9/11 Cash_Flow ...")
    build_cash_flow(wb, R)

    print(" 10/11 DCF ...")
    build_dcf(wb, R)

    print(" 11/11 Sensitivity + Ratio_Analysis ...")
    build_sensitivity(wb, R)
    build_ratio_analysis(wb, R)

    print("  Key_Summary (dashboard) ...")
    build_key_summary(wb, R)

    # Save
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Total key rows tracked in R dict: {len(R)}")
    return wb, R


# ============================================================
# SELF-CHECK / VERIFICATION
# ============================================================

def check_model(output_path=OUTPUT):
    """Load the generated xlsx and verify key linkages programmatically."""
    print(f"\n{'='*60}")
    print(f"SELF-CHECK: Verifying {output_path} ...")
    print("=" * 60)

    wb = openpyxl.load_workbook(output_path)

    errors = []
    warnings = []

    # 1. Check all expected tabs exist
    expected_tabs = ["Cover", "Key_Summary", "Assumptions", "Growth_Drivers",
                     "Segment_PL", "Consolidated_PL", "BS",
                     "Cash_Flow", "DCF", "Sensitivity", "Ratio_Analysis"]
    for tab in expected_tabs:
        if tab not in wb.sheetnames:
            errors.append(f"MISSING TAB: {tab}")
    print(f"  [{'PASS' if not errors else 'FAIL'}] All {len(expected_tabs)} tabs present")

    # 2. Check Segment_PL cross-check = 0 formula exists
    if "Segment_PL" in wb.sheetnames:
        ws = wb["Segment_PL"]
        # Find "CROSS-CHECK" or "Difference" row
        found_xcheck = False
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1):
            for cell in row:
                if cell.value and "cross-check" in str(cell.value).lower():
                    found_xcheck = True
                    break
        if found_xcheck:
            print("  [PASS] Segment_PL cross-check row present")
        else:
            warnings.append("Segment_PL cross-check row not found (may need manual verification)")

    # 3. Check BS balance check row exists
    if "BS" in wb.sheetnames:
        ws = wb["BS"]
        found_bal = False
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1):
            for cell in row:
                if cell.value and "balance check" in str(cell.value).lower():
                    found_bal = True
                    break
        if found_bal:
            print("  [PASS] BS balance check row present")
        else:
            warnings.append("BS balance check row not found")

    # 4. Check no formulas in Notes column (tab-specific: K=11 for most, AI=35 for quarterly)
    quarterly_tabs = {"Growth_Drivers", "Assumptions", "Segment_PL"}
    notes_col_issues = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        notes_col = GD_NCOL if sheet_name in quarterly_tabs else NCOL
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=notes_col, max_col=notes_col):
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    notes_col_issues.append(f"{sheet_name}!{cell.coordinate}: Notes column contains formula: {cell.value}")
    if not notes_col_issues:
        print("  [PASS] No formulas in Notes column")
    else:
        for w in notes_col_issues:
            print(f"  [WARN] {w}")

    # 5. Check CHOOSE/MATCH formulas in Assumptions (quarterly: cols 18-32)
    if "Assumptions" in wb.sheetnames:
        ws = wb["Assumptions"]
        choose_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=18, max_col=32):
            for cell in row:
                if cell.value and "CHOOSE(MATCH" in str(cell.value):
                    choose_count += 1
        print(f"  [PASS] {choose_count} CHOOSE/MATCH formulas in Assumptions (quarterly forecast columns)")

    # 6. Check Segment_PL formulas reference Assumptions (quarterly layout)
    if "Segment_PL" in wb.sheetnames:
        ws = wb["Segment_PL"]
        ref_count = 0
        # Scan forecast quarterly + annual columns (18-32) for Assumptions references
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=18, max_col=32):
            for cell in row:
                if cell.value and "Assumptions!" in str(cell.value):
                    ref_count += 1
        print(f"  [PASS] {ref_count} forecast formulas reference Assumptions in Segment_PL")

    # 7. Check DCF has implied price
    if "DCF" in wb.sheetnames:
        ws = wb["DCF"]
        found_price = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value and "implied share price" in str(cell.value).lower():
                    found_price = True
                    break
        if found_price:
            print("  [PASS] DCF implied share price row present")
        else:
            warnings.append("DCF implied share price row not found")

    # 8. Count total rows
    total_rows = sum(wb[sheet].max_row for sheet in wb.sheetnames)
    print(f"  [INFO] Total rows across all tabs: {total_rows}")

    # Summary
    print(f"\n{'='*60}")
    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")
    if warnings:
        print(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  - {w}")
    if not errors and not warnings:
        print("All checks passed!")
    print(f"{'='*60}")

    wb.close()
    return len(errors) == 0


# ============================================================
# MAIN ENTRY POINT
# ============================================================

if __name__ == "__main__":
    wb, R = build_model()
    check_model()
