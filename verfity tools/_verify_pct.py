import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model_v2.xlsx')

# Check Assumptions share values
ws_a = wb['Assumptions']
print('=== Assumptions - Share & % values (should be 0.x, not xx.0) ===')
for row in range(1, ws_a.max_row + 1):
    a = ws_a.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).replace('−', '-')
    if any(kw in a_str for kw in ['Intel Share', 'Intel DC Share', 'D/(D+E)', 'Risk-Free',
                                     'Cost of Debt', 'Tax Rate', 'R&D %', 'MG&A %', 'COGS',
                                     'Accrued Comp', 'Other CA', 'Other CL', 'Income Tax',
                                     'Terminal Growth']):
        vals = [ws_a.cell(row=row, column=c).value for c in [2,3,4,6,7,8,9]]
        print(f'Row {row} [{a_str[:60]}]: {vals}')

# Check Growth_Drivers - CCG implied revenue (should be reasonable now)
ws_g = wb['Growth_Drivers']
print('\n=== Growth_Drivers - CCG Implied Revenue ===')
for row in range(1, ws_g.max_row + 1):
    a = ws_g.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).replace('−', '-')
    if 'Implied CCG' in a_str:
        vals = [ws_g.cell(row=row, column=c).value for c in [6,7,8]]
        print(f'Row {row} [{a_str[:60]}]: FY26E={vals[0]} | FY27E={vals[1]} | FY28E={vals[2]}')
    if 'Implied DCAI' in a_str:
        vals = [ws_g.cell(row=row, column=c).value for c in [6,7,8]]
        print(f'Row {row} [{a_str[:60]}]: FY26E={vals[0]} | FY27E={vals[1]} | FY28E={vals[2]}')

wb.close()
