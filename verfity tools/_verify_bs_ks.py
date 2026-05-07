import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model.xlsx')

# Check BS balance check
ws_bs = wb['BS']
print('=== BS BALANCE CHECK ===')
for row in range(1, ws_bs.max_row + 1):
    a = ws_bs.cell(row=row, column=1).value
    if a and 'balance' in str(a).lower():
        print(f"Found 'balance' at Row {row}: {str(a)[:60]}")
        for check_row in range(row-1, min(row+5, ws_bs.max_row+1)):
            a2 = ws_bs.cell(row=check_row, column=1).value
            f6 = ws_bs.cell(row=check_row, column=6).value
            f8 = ws_bs.cell(row=check_row, column=8).value
            print(f"  Row {check_row} [{str(a2)[:50] if a2 else 'empty'}]: F6={str(f6)[:100] if f6 else 'empty'}, H8={str(f8)[:100] if f8 else 'empty'}")

# Also check total assets and L&E
for row in range(1, ws_bs.max_row + 1):
    a = ws_bs.cell(row=row, column=1).value
    if a and ('TOTAL ASSET' in str(a).upper() or 'TOTAL LIABILIT' in str(a).upper() or "TOTAL STOCKHOLDERS" in str(a).upper()):
        for c in [2,3,4,6,7,8]:
            v = ws_bs.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  Row {row} [{str(a)[:60]}]: {cl}={str(v)[:100] if v else 'empty'}")

# Check Key_Summary revenue links
ws_ks = wb['Key_Summary']
print()
print('=== KEY_SUMMARY Revenue ===')
for row in range(1, 50):
    a = ws_ks.cell(row=row, column=1).value
    if a and 'revenue' in str(a).lower():
        for c in [2,3,4,6,7,8]:
            v = ws_ks.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  Row {row} [{str(a)[:50]}]: {cl}={str(v)[:100] if v else 'N/A'}")

wb.close()
