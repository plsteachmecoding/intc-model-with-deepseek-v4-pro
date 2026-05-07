import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model.xlsx')

# Check Segment_Revenue consolidated
ws = wb['Segment_Revenue']
print('=== SEGMENT_REVENUE ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and any(kw in str(a).upper() for kw in ['TOTAL CONSOLIDATED','EXTERNAL FOUNDRY','INTERSEGMENT','TOTAL SEGMENT','CCG -','DCAI -','INTEL FOUNDRY','ALL OTHER']):
        vals = {}
        for c in [2,3,4,6,7,8]:
            v = ws.cell(row=row, column=c).value
            vals[c] = str(v)[:110] if v else '(empty)'
            col_l = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
        print(f"Row {row} [{str(a)[:60]}]: B={vals[2]} | C={vals[3]} | D={vals[4]} | F={vals[6]} | G={vals[7]} | H={vals[8]}")

# Check Consolidated_PL revenue row
ws_pl = wb['Consolidated_PL']
print()
print('=== CONSOLIDATED_PL Revenue ===')
for row in range(1, 40):
    a = ws_pl.cell(row=row, column=1).value
    if a and ('revenue' in str(a).lower() or 'total' in str(a).lower()):
        for c in [2,3,4,6,7,8]:
            v = ws_pl.cell(row=row, column=c).value
            col_l = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  Row {row} [{str(a)[:50]}]: {col_l}={str(v)[:110] if v else 'N/A'}")

# Check Segment_PL Foundry OI
ws_sp = wb['Segment_PL']
print()
print('=== SEGMENT_PL Foundry ===')
for row in range(1, ws_sp.max_row + 1):
    a = ws_sp.cell(row=row, column=1).value
    if a and 'foundry' in str(a).lower():
        for c in [2,3,4,6,7,8]:
            v = ws_sp.cell(row=row, column=c).value
            col_l = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  Row {row} [{str(a)[:50]}]: {col_l}={str(v)[:110] if v else 'N/A'}")

wb.close()
