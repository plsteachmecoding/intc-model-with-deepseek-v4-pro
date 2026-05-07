import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model_v2.xlsx')
ws = wb['Growth_Drivers']
print('=== Driver Sensitivity Section ===')
in_sens = False
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and 'Driver Sensitivity' in str(a):
        in_sens = True
    if not in_sens:
        continue
    if a is None:
        continue
    a_str = str(a).replace('−', '-')
    vals = {}
    for c in [6,7,8,9]:
        v = ws.cell(row=row, column=c).value
        vals[c] = str(v)[:80] if v else '-'
    print(f'Row {row} [{a_str[:55]}]: F={vals.get(6,"?")} | G={vals.get(7,"?")} | H={vals.get(8,"?")} | I={vals.get(9,"?")}')
    if a and 'Top 3' in str(a):
        break
wb.close()
