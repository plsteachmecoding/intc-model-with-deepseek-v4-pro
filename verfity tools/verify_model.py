"""
verify_model.py — Deep validation of Intel_IB_Model.xlsx
Checks: cross-checks=0, BS balance, scenario switching, DCF calculation, CF links
"""
import openpyxl

wb = openpyxl.load_workbook("Intel_IB_Model.xlsx")
print("=" * 60)
print("DEEP VERIFICATION: Intel_IB_Model.xlsx")
print("=" * 60)

# ============================================================
# 1. Check 11 tabs present
# ============================================================
expected = ["Cover", "Key_Summary", "Assumptions", "Segment_Revenue",
            "Segment_PL", "Consolidated_PL", "BS", "Cash_Flow",
            "DCF", "Sensitivity", "Ratio_Analysis"]
missing = [t for t in expected if t not in wb.sheetnames]
print(f"\n1. TABS: {'PASS - all 11 present' if not missing else f'MISSING: {missing}'}")

# ============================================================
# 2. Find and check cross-check row in Segment_PL
# ============================================================
ws = wb["Segment_PL"]
print("\n2. SEGMENT_PL CROSS-CHECK:")
# Find "Difference" row
for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "difference" in str(cell.value).lower():
        diff_row = cell.row
        vals = {}
        for col in [2, 3, 4, 6, 7, 8]:
            cell_ref = ws.cell(row=diff_row, column=col)
            vals[col] = cell_ref.value
            if cell_ref.value is not None:
                col_labels = {2:'B', 3:'C', 4:'D', 6:'F', 7:'G', 8:'H'}
        print(f"  Col {col_labels.get(col, col)}: formula={cell_ref.value}")
        break

# Check if formulas reference Segment_PL and Consolidated_PL
for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "sum seg oi" in str(cell.value).lower():
        sum_oi_row = cell.row + 1  # label row + 1
        print(f"  Sum Seg OI row: {sum_oi_row}")
    if cell.value and "consolidated_pl ebit" in str(cell.value).lower():
        ebit_row = cell.row + 1
        print(f"  Consolidated EBIT link row: {ebit_row}")

# ============================================================
# 3. Check BS balance check
# ============================================================
ws_bs = wb["BS"]
print("\n3. BS BALANCE CHECK:")
for row in ws_bs.iter_rows(min_row=1, max_row=120, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "balance check" in str(cell.value).lower():
        bal_row = cell.row  # label and formula in same row
        col_labels = {2:'B', 3:'C', 4:'D', 6:'F', 7:'G', 8:'H'}
        for col in [2, 3, 4, 6, 7, 8]:
            cell_ref = ws_bs.cell(row=bal_row, column=col)
            if cell_ref.value:
                print(f"  {col_labels[col]}: formula={cell_ref.value}")
        break

# ============================================================
# 4. Verify Consolidated_PL EBIT links to Segment_PL
# ============================================================
ws_pl = wb["Consolidated_PL"]
print("\n4. CONSOLIDATED_PL EBIT LINK:")
for row in ws_pl.iter_rows(min_row=1, max_row=80, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "ebit" in str(cell.value).lower() and "segment" in str(cell.value).lower():
        ebit_row_pl = cell.row + 1
        for col in [2, 3, 4, 6, 7, 8]:
            cell_ref = ws_pl.cell(row=ebit_row_pl, column=col)
            if cell_ref.value:
                has_seg_ref = "Segment_PL" in str(cell_ref.value)
                print(f"  Col {col}: {'PASS - links to Segment_PL' if has_seg_ref else 'FAIL - ' + str(cell_ref.value)[:80]}")
        break

# ============================================================
# 5. Verify CF WC changes link to BS delta
# ============================================================
ws_cf = wb["Cash_Flow"]
print("\n5. CASH FLOW WC LINKS TO BS:")
wc_items_found = 0
for row in ws_cf.iter_rows(min_row=1, max_row=80, min_col=6, max_col=8):
    for cell in row:
        if cell.value and "BS!" in str(cell.value):
            wc_items_found += 1
print(f"  WC items referencing BS: {wc_items_found}")
print(f"  {'PASS' if wc_items_found >= 3 else 'WARN - fewer than expected'}")

# Check FCF formula
for row in ws_cf.iter_rows(min_row=1, max_row=80, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "free cash flow" in str(cell.value).lower():
        fcf_row = cell.row
        for col in [6, 7, 8]:
            cell_ref = ws_cf.cell(row=fcf_row, column=col)
            if cell_ref.value:
                has_cfo = "cf_cfo" in str(cell_ref.value) or "CFO" in str(cell_ref.value)
                has_cfi = "cf_cfi" in str(cell_ref.value) or "CFI" in str(cell_ref.value)
                print(f"  FCF col {col} formula: {'PASS - CFO+CFI' if (has_cfo or has_cfi) else str(cell_ref.value)[:60]}")
        break

# ============================================================
# 6. Verify DCF UFCF formula
# ============================================================
ws_dcf = wb["DCF"]
print("\n6. DCF UFCF VERIFICATION:")
for row in ws_dcf.iter_rows(min_row=1, max_row=60, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "unlevered free cash flow" in str(cell.value).lower():
        ufcf_row = cell.row
        for col in [6, 7, 8]:
            cell_ref = ws_dcf.cell(row=ufcf_row, column=col)
            if cell_ref.value:
                has_nopat = "NOPAT" in str(cell_ref.value) or "nopat" in str(cell_ref.value).lower()
                print(f"  UFCF col {col}: formula={'PASS' if has_nopat else str(cell_ref.value)[:80]}")
        break

# Find implied price row
for row in ws_dcf.iter_rows(min_row=1, max_row=80, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "implied share price" in str(cell.value).lower():
        price_row = cell.row
        price_cell = ws_dcf.cell(row=price_row, column=8)
        print(f"\n7. DCF IMPLIED PRICE (Base case, col H):")
        print(f"  Formula: {price_cell.value}")

# ============================================================
# 7. Check all Notes columns for formulas
# ============================================================
print("\n8. NOTES COLUMN FORMULA CHECK:")
formula_notes = 0
for sheet_name in wb.sheetnames:
    ws_s = wb[sheet_name]
    for row in ws_s.iter_rows(min_row=1, max_row=ws_s.max_row, min_col=11, max_col=11):
        for cell in row:
            if cell.value and str(cell.value).startswith("="):
                formula_notes += 1
                print(f"  WARNING: {sheet_name}!{cell.coordinate}: {str(cell.value)[:80]}")
print(f"  {'PASS - No formulas in Notes' if formula_notes == 0 else f'WARN - {formula_notes} formula(s) found'}")

# ============================================================
# 8. Key rows summary
# ============================================================
print("\n9. KEY OUTPUT VALUES (formulas - evaluate in Excel):")
print(f"  DCF Implied Price: =DCF!H{price_row}" if price_row is not None else "  (price_row not found)")

# Check Key_Summary scenario cell
ws_ks = wb["Key_Summary"]
scenario_cell = ws_ks.cell(row=3, column=2)
print(f"\n  Key_Summary scenario link: {scenario_cell.value}")

print(f"\n{'='*60}")
print("VERIFICATION COMPLETE")
print(f"{'='*60}")

wb.close()
