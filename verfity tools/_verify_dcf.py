import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model_v2.xlsx')

# DCF structure
ws = wb['DCF']
print('=== DCF Full Structure ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).strip().replace(chr(0x2212), '-')
    vals = []
    for c in [5, 6, 7, 8, 9]:
        v = ws.cell(row=row, column=c).value
        if v is not None:
            vals.append((c, str(v)[:55]))
    has_content = any(ws.cell(row=row, column=c).value is not None for c in [6, 7, 8, 9])
    if has_content and vals:
        parts = ["col{}={}".format(c, s) for c, s in vals]
        print("Row {} [{}]: {}".format(row, a_str[:55], " | ".join(parts[:4])))

print()
print('=== Consolidated P&L Key Lines ===')
ws2 = wb['Consolidated_PL']
for row in range(1, ws2.max_row + 1):
    a = ws2.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).strip()
    if any(kw in a_str for kw in ['Total Revenue', 'EBIT', 'Net Income', 'Diluted EPS', 'Gross Profit', 'Operating Income', 'Income Before Tax']):
        vals = []
        for c in [5, 6, 7, 8, 9]:
            v = ws2.cell(row=row, column=c).value
            vals.append(str(v)[:35] if v is not None else '-')
        if any(v != '-' for v in vals):
            print("Row {} [{}]: F={} | G={} | H={} | I={}".format(row, a_str[:40], vals[1], vals[2], vals[3], vals[4]))

print()
print('=== BS Key Lines ===')
ws3 = wb['BS']
for row in range(1, ws3.max_row + 1):
    a = ws3.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).strip()
    if any(kw in a_str for kw in ['Total Assets', 'Total L&E', 'Cash & Equivalents', 'Total Equity', 'BALANCE CHECK']):
        vals = []
        for c in [5, 6, 7, 8, 9]:
            v = ws3.cell(row=row, column=c).value
            vals.append(str(v)[:35] if v is not None else '-')
        print("Row {} [{}]: F={} | G={} | H={} | I={}".format(row, a_str[:40], vals[1], vals[2], vals[3], vals[4]))

print()
print('=== Ratio Analysis Key Metrics ===')
ws4 = wb['Ratio_Analysis']
for row in range(1, min(ws4.max_row + 1, 60)):
    a = ws4.cell(row=row, column=1).value
    if a is None: continue
    a_str = str(a).strip()
    if any(kw in a_str for kw in ['GPM', 'OPM', 'NPM', 'ROE', 'ROA', 'D/A', 'Implied', 'FCF/Rev']):
        vals = []
        for c in [5, 6, 7, 8, 9]:
            v = ws4.cell(row=row, column=c).value
            vals.append(str(v)[:30] if v is not None else '-')
        print("Row {} [{}]: F={} | G={} | H={} | I={}".format(row, a_str[:40], vals[1], vals[2], vals[3], vals[4]))

wb.close()
