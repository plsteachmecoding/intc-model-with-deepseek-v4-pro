import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model.xlsx')
ws = wb['Ratio_Analysis']
print('=== VALUATION MULTIPLES ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and any(kw in str(a).upper() for kw in ['STOCK PRICE','MARKET VALUATION','TTM VALUATION','NTM VALUATION','P/E','P/B','P/S','EV/EBITDA','EV/REVENUE','P/CF','PEG','DIVIDEND YIELD']):
        d_val = ws.cell(row=row, column=4).value
        f_val = ws.cell(row=row, column=6).value
        print(f"Row {row} [{str(a)[:55]}]: D={str(d_val)[:130] if d_val else '-'} | F={str(f_val)[:130] if f_val else '-'}")
wb.close()
