import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model_v2.xlsx')
ws = wb['Growth_Drivers']
print('=== GROWTH_DRIVERS - DCAI SECTION ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a is None:
        continue
    a_str = str(a).replace('−', '-')
    kw_list = ['DCAI:', 'Server TAM', 'Intel DC Server', 'Intel Implied DC',
               'Intel Server Blended', 'Implied CPU', 'AI/ASIC', 'Implied DCAI',
               'Reported DCAI', 'Reconciliation:', 'Reconcil. as %',
               'Method: Revenue = Server']
    match = any(kw in a_str for kw in kw_list)
    if match:
        vals = []
        for c in [2,3,4,6,7,8]:
            v = ws.cell(row=row, column=c).value
            vals.append(str(v)[:100] if v else '-')
        print(f"Row {row} [{a_str[:80]}]: {vals}")
wb.close()
