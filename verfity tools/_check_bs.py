import openpyxl
wb = openpyxl.load_workbook("Intel_IB_Model.xlsx")
ws = wb["BS"]
print("--- BS KEY ROWS ---")
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and ("balance" in str(a).lower() or "plug" in str(a).lower() or "total asset" in str(a).lower() or "total liab" in str(a).lower()):
        vals = {}
        for c in [2, 3, 4, 6, 7, 8]:
            v = ws.cell(row=row, column=c).value
            vals[c] = str(v)[:80] if v else "(empty)"
        print(f"Row {row} [{str(a)[:60]}]: B={vals[2]} | F={vals[6]} | H={vals[8]}")
print()

# Check specific rows around cash plug and balance
print("--- CASH PLUG (row 11-13) ---")
for row in range(10, 15):
    a = ws.cell(row=row, column=1).value
    f6 = ws.cell(row=row, column=6).value
    print(f"Row {row} [{str(a)[:50] if a else 'empty'}]: F={str(f6)[:100] if f6 else 'empty'}")

print()
print("--- TOTAL ASSETS + TOTAL L&E ---")
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and ("TOTAL ASSET" in str(a).upper() or "TOTAL LIABILIT" in str(a).upper() or "TOTAL STOCKHOLDERS" in str(a).upper()):
        vals = {}
        for c in [2, 3, 4, 6, 7, 8]:
            v = ws.cell(row=row, column=c).value
            vals[c] = str(v)[:100] if v else "(empty)"
        print(f"Row {row} [{str(a)[:60]}]: F={vals[6]} | H={vals[8]}")

# Balance check
print()
print("--- BALANCE CHECK ---")
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and "balance check" in str(a).lower():
        for check_row in range(row, row + 5):
            a2 = ws.cell(row=check_row, column=1).value
            f6 = ws.cell(row=check_row, column=6).value
            f8 = ws.cell(row=check_row, column=8).value
            print(f"  Row {check_row} [{str(a2)[:50] if a2 else 'empty'}]: F6={str(f6)[:100] if f6 else 'empty'}, H8={str(f8)[:100] if f8 else 'empty'}")

wb.close()
