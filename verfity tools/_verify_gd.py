import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model_v2.xlsx')
ws = wb['Growth_Drivers']
print('=== GROWTH_DRIVERS — CCG SECTION ===')
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a and any(kw in str(a).upper() for kw in ['CCG', 'PC TAM', 'INTEL PC', 'INTEL IMPLIED', 'IMPLIED CCG', 'REPORTED CCG', 'RECONCIL', 'BLENDED ASP', 'BOTTOM-UP', 'METHOD']):
        vals = {}
        for c in [2,3,4,6,7,8]:
            v = ws.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            vals[cl] = str(v)[:130] if v else '-'
        print(f"Row {row} [{str(a)[:65]}]: D={vals.get('D','?')} | F={vals.get('F','?')}")
wb.close()
