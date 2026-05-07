import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model.xlsx')
ws_bs = wb['BS']

# Find actual "Balance Check" label (not title)
print('=== BS BALANCE CHECK ROW ===')
for row in range(3, ws_bs.max_row + 1):
    a = ws_bs.cell(row=row, column=1).value
    if a and 'balance check' in str(a).lower() and 'intel' not in str(a).lower():
        print(f"Found at Row {row}: [{str(a)}]")
        for c in [2,3,4,6,7,8]:
            v = ws_bs.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  {cl}={str(v)[:150] if v else 'empty'}")

# Also show all rows with "check" in name
print()
print('=== ALL ROWS WITH "CHECK" ===')
for row in range(1, ws_bs.max_row + 1):
    a = ws_bs.cell(row=row, column=1).value
    if a and 'check' in str(a).lower():
        print(f"Row {row}: [{str(a)[:60]}]")

# Show total assets, total L&E, equity rows
print()
print('=== KEY BS TOTALS ===')
for row in range(1, ws_bs.max_row + 1):
    a = ws_bs.cell(row=row, column=1).value
    if a and any(kw in str(a).upper() for kw in ['TOTAL ASSET', 'TOTAL LIABILIT', 'TOTAL EQUITY', 'TOTAL STOCKHOLDER', 'TOTAL L&E', 'BALANCE CHECK']):
        for c in [2,3,4,6,7,8]:
            v = ws_bs.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"  Row {row} [{str(a)[:60]}]: {cl}={str(v)[:120] if v else 'empty'}")

# Show cash plug row
print()
print('=== CASH PLUG ROW ===')
for row in range(10, 25):
    a = ws_bs.cell(row=row, column=1).value
    f6 = ws_bs.cell(row=row, column=6).value
    print(f"Row {row} [{str(a)[:50] if a else 'empty'}]: F6={str(f6)[:150] if f6 else 'empty'}")

wb.close()
