import openpyxl
wb = openpyxl.load_workbook('Intel_IB_Model.xlsx')
ws = wb['BS']
print('=== CASH & EQUIVALENTS ===')
for row in range(5, 15):
    a = ws.cell(row=row, column=1).value
    if a:
        for c in [2,3,4,6,7,8]:
            v = ws.cell(row=row, column=c).value
            cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
            print(f"Row {row} [{str(a)[:50]}]: {cl}={str(v)[:200] if v else 'empty'}")

print()
print('=== TOTAL L&E (Row 43) ===')
for c in [2,3,4,6,7,8]:
    v = ws.cell(row=43, column=c).value
    cl = {2:'B',3:'C',4:'D',6:'F',7:'G',8:'H'}[c]
    print(f"{cl}={str(v)[:200] if v else 'empty'}")

wb.close()
